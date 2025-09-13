#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据对比器 - 用于对比Excel数据与数据库数据
"""

import sys
import os
from typing import Dict, List, Any, Optional, Tuple
from loguru import logger
from datetime import datetime, timedelta
import re
from difflib import SequenceMatcher
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 添加项目根目录到路径
project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.append(project_root)

from services.database_manager import get_db_manager
from services.config_database_manager import config_db_manager
from backend.constants.field_config import FIELD_CONFIG, EXCEL_METRICS_MAPPING, get_field_display_name


class DataComparator:
    """数据对比器类"""
    
    def __init__(self, target_month: str = None, excel_filename: str = None):
        """
        初始化数据对比器
        
        Args:
            target_month: 目标月份，格式：2025-09
            excel_filename: Excel文件名，用于提取月份信息
        """
        self.excel_filename = excel_filename
        
        # 如果提供了Excel文件名，尝试从中提取月份
        if excel_filename and not target_month:
            target_month = self._extract_month_from_filename(excel_filename)
        
        self.target_month = target_month
        self.db_manager = None
        self.errors = []
        self.warnings = []
        self.comparison_data = {}
        
        # 初始化数据库连接
        self.db_manager = get_db_manager()
    
    def _extract_month_from_filename(self, filename: str) -> str:
        """
        从Excel文件名中提取月份信息
        
        Args:
            filename: Excel文件名
            
        Returns:
            月份字符串，格式：2025-09，如果无法提取则返回None
        """
        try:
            import re
            
            # 常见的月份模式
            patterns = [
                r'(\d{4})[年\-_](\d{1,2})[月\-_]',  # 2025年9月, 2025-09, 2025_09
                r'(\d{4})(\d{2})',  # 202509
                r'(\d{1,2})[月\-_](\d{4})',  # 9月2025, 09-2025
                r'[（(](\d{1,2})月[）)]',  # （8月）或 (8月)
            ]
            
            for pattern in patterns:
                match = re.search(pattern, filename)
                if match:
                    if len(match.groups()) == 2:
                        year, month = match.groups()
                        # 确保月份是两位数
                        month = month.zfill(2)
                        return f"{year}-{month}"
                    elif len(match.groups()) == 1:
                        # 只有月份的情况，使用当前年份
                        month = match.groups()[0]
                        month = month.zfill(2)
                        from datetime import datetime
                        current_year = datetime.now().year
                        return f"{current_year}-{month}"
            
            # 如果无法从文件名提取，尝试从当前日期推断
            logger.warning(f"无法从文件名 '{filename}' 中提取月份信息，使用当前月份")
            from datetime import datetime
            current_date = datetime.now()
            return f"{current_date.year}-{current_date.month:02d}"
            
        except Exception as e:
            logger.error(f"提取月份信息异常: {e}")
            return None
    
    async def compare_data(self, excel_data: Dict[str, Any], stores: List[Dict]) -> Dict[str, Any]:
        """
        对比Excel数据与数据库数据
        
        Args:
            excel_data: Excel中读取的数据
            stores: 门店列表
            
        Returns:
            Dict: 对比结果
        """
        try:
            logger.info(f"开始数据对比，Excel门店数量: {len(excel_data)}, 预期门店数量: {len(stores)}")
            
            # 重置错误和警告
            self.errors = []
            self.warnings = []
            self.comparison_data = {}
            
            # 检查数据库文件覆盖情况
            database_info = await self._check_database_coverage(excel_data)
            
            # 获取字段配置
            field_rule = config_db_manager.get_export_field_rule()
            if field_rule and field_rule.get('selected_fields'):
                selected_fields = field_rule['selected_fields']
            else:
                selected_fields = list(FIELD_CONFIG.keys())
            
            logger.info(f"使用字段配置: {len(selected_fields)} 个字段")
            
            # 为每个Excel门店查找匹配的数据库门店
            for excel_store_name, excel_store_data in excel_data.items():
                logger.info(f"处理Excel门店: {excel_store_name}")
                
                # 查找匹配的数据库门店
                matched_stores = await self._find_matching_stores(excel_store_name)
                
                if len(matched_stores) == 0:
                    # 未找到匹配门店
                    self.errors.append({
                        "type": "store_not_found",
                        "excel_store": excel_store_name,
                        "message": f"未找到匹配的门店: {excel_store_name}"
                    })
                    continue
                elif len(matched_stores) > 1:
                    # 找到多个匹配门店
                    self.errors.append({
                        "type": "multiple_stores_found",
                        "excel_store": excel_store_name,
                        "matched_stores": [store['storeName'] for store in matched_stores],
                        "message": f"门店 {excel_store_name} 匹配到多个数据库门店: {', '.join([store['storeName'] for store in matched_stores])}"
                    })
                    continue
                
                # 找到唯一匹配门店
                db_store = matched_stores[0]
                logger.info(f"Excel门店 '{excel_store_name}' 匹配到数据库门店 '{db_store['storeName']}'")
                
                # 对比该门店的数据
                store_comparison = await self._compare_store_data(
                    excel_store_name, 
                    excel_store_data, 
                    db_store, 
                    selected_fields
                )
                
                self.comparison_data[excel_store_name] = store_comparison
            
            # 检查是否有错误
            has_errors = len(self.errors) > 0
            
            # 记录返回给前端的数据库信息
            logger.info(f"返回给前端的数据库信息: coverage_rate={database_info.get('coverage_rate', 'N/A')}, "
                       f"excel_dates_count={database_info.get('excel_dates_count', 'N/A')}, "
                       f"available_dates_count={database_info.get('available_dates_count', 'N/A')}")
            
            return {
                "has_errors": has_errors,
                "errors": self.errors,
                "warnings": self.warnings,
                "comparison_data": self.comparison_data,
                "database_info": database_info
            }
            
        except Exception as e:
            logger.error(f"数据对比异常: {e}")
            raise e
    
    async def _find_matching_stores(self, excel_store_name: str) -> List[Dict]:
        """
        查找匹配的数据库门店
        
        Args:
            excel_store_name: Excel中的门店名称
            
        Returns:
            List[Dict]: 匹配的门店列表
        """
        try:
            if not self.db_manager:
                return []
            
            # 获取所有门店名称（从数据库中查询唯一的MDShow值）
            all_stores = self._get_all_unique_stores()
            matched_stores = []
            
            logger.info(f"开始匹配Excel门店: '{excel_store_name}'")
            logger.info(f"数据库中共有 {len(all_stores)} 个门店")
            
            # 清理Excel门店名称
            clean_excel_name = self._clean_store_name(excel_store_name)
            logger.info(f"清理后的Excel门店名: '{clean_excel_name}'")
            
            # 优化匹配策略：先尝试精确搜索，再尝试模糊搜索
            # 策略1: 精确匹配（直接搜索）
            exact_matches = [store for store in all_stores if self._clean_store_name(store) == clean_excel_name]
            if exact_matches:
                for store_name in exact_matches:
                    matched_stores.append({
                        'storeName': self._clean_display_name(store_name),
                        'id': store_name
                    })
                    logger.info(f"精确匹配: '{excel_store_name}' -> '{store_name}'")
            
            # 策略2: 包含匹配（搜索包含关系）
            if not matched_stores:
                for store_name in all_stores:
                    clean_db_name = self._clean_store_name(store_name)
                    if clean_excel_name in clean_db_name or clean_db_name in clean_excel_name:
                        matched_stores.append({
                            'storeName': self._clean_display_name(store_name),
                            'id': store_name
                        })
                        logger.info(f"包含匹配: '{excel_store_name}' -> '{store_name}'")
                        break  # 找到第一个包含匹配就停止
            
            # 策略3: 相似度匹配（只在前面策略都失败时使用）
            if not matched_stores:
                best_similarity = 0
                best_match = None
                
                for store_name in all_stores:
                    clean_db_name = self._clean_store_name(store_name)
                    similarity = SequenceMatcher(None, clean_excel_name, clean_db_name).ratio()
                    
                    if similarity > best_similarity and similarity > 0.6:
                        best_similarity = similarity
                        best_match = store_name
                
                if best_match:
                    matched_stores.append({
                        'storeName': self._clean_display_name(best_match),
                        'id': best_match
                    })
                    logger.info(f"相似度匹配: '{excel_store_name}' -> '{best_match}' (相似度: {best_similarity:.3f})")
            
            if not matched_stores:
                logger.warning(f"未找到匹配门店: '{excel_store_name}' (清理后: '{clean_excel_name}')")
                # 输出前几个数据库门店名称作为参考
                sample_stores = all_stores[:5]
                logger.info(f"数据库门店示例: {sample_stores}")
            
            return matched_stores
            
        except Exception as e:
            logger.error(f"查找匹配门店异常: {e}")
            return []
    
    def _get_all_unique_stores(self) -> List[str]:
        """获取数据库中所有唯一的门店名称"""
        try:
            if not self.db_manager:
                return []
            
            # 使用数据库管理器的get_stores方法
            stores_data = self.db_manager.get_stores()
            
            # 提取门店名称
            store_names = []
            for store in stores_data:
                name = store.get('name') or store.get('value') or store.get('storeName', '')
                if name and name not in store_names:
                    store_names.append(name)
            
            return store_names
            
        except Exception as e:
            logger.error(f"获取门店列表异常: {e}")
            return []
    
    def _clean_store_name(self, name: str) -> str:
        """清理门店名称，用于匹配"""
        if not name:
            return ""
        
        # 移除HTML标签
        name = re.sub(r'<[^>]+>', '', name)
        
        # 移除常见的前缀（更精确的模式）
        name = re.sub(r'^\[[^\]]*\]\s*', '', name)  # 移除 [S019] 这样的前缀
        name = re.sub(r'^[A-Z]\d+\s*', '', name)   # 移除 S001 这样的前缀
        
        # 移除括号内容（但保留主要内容）
        name = re.sub(r'\s*\([^)]*\)$', '', name)  # 移除末尾括号
        
        # 标准化空格（保留基本空格，只合并多个空格）
        name = re.sub(r'\s+', ' ', name)
        
        # 移除特殊字符但保留中文、字母、数字
        name = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9\s]', '', name)
        
        return name.strip()
    
    def _clean_display_name(self, name: str) -> str:
        """清理门店名称用于显示，移除HTML标签和格式化字符"""
        if not name:
            return ""
        
        # 移除HTML标签
        name = re.sub(r'<[^>]+>', '', name)
        
        # 移除常见的前缀（更精确的模式）
        name = re.sub(r'^\[[^\]]*\]\s*', '', name)  # 移除 [S019] 这样的前缀
        name = re.sub(r'^[A-Z]\d+\s*', '', name)   # 移除 S001 这样的前缀
        
        # 移除括号内容（但保留主要内容）
        name = re.sub(r'\s*\([^)]*\)$', '', name)  # 移除末尾括号
        
        # 标准化空格
        name = re.sub(r'\s+', ' ', name)
        
        return name.strip()
    
    def _convert_to_excel_value(self, value: Any) -> Any:
        """将值转换为Excel兼容的类型"""
        if value is None:
            return ""
        
        # 如果是数字，检查是否需要转换为百分比格式
        if isinstance(value, (int, float)):
            # 如果是0-1之间的小数，很可能是数据库中的百分比数据，转换为百分比格式
            if 0 <= value <= 1:
                return f"{round(value * 100, 2)}%"
            else:
                return value
        
        # 如果已经是字符串且包含百分号，直接返回
        if isinstance(value, str) and '%' in value:
            return value
        
        # 如果是字典，检查是否是错误的数据结构
        if isinstance(value, dict):
            logger.warning(f"检测到字典类型的值: {value}，这可能表示数据结构有问题")
            # 如果字典只有一个值，尝试返回这个值
            if len(value) == 1:
                return list(value.values())[0]
            return str(value)
        
        # 如果是列表，转换为字符串
        if isinstance(value, list):
            return str(value)
        
        # 其他类型转换为字符串
        return str(value)
    
    async def _check_database_coverage(self, excel_data: Dict[str, Any]) -> Dict[str, Any]:
        """检查数据库文件覆盖情况"""
        try:
            if not self.target_month:
                return {"has_coverage_issues": False, "message": "未指定目标月份"}
            
            # 获取Excel中的所有日期
            excel_dates = set()
            for store_name, store_data in excel_data.items():
                logger.debug(f"检查门店 {store_name} 的数据结构")
                for metric_name, metric_data in store_data.items():
                    logger.debug(f"指标 {metric_name}, 数据类型: {type(metric_data)}, 值: {str(metric_data)[:100]}")
                    
                    # 检查metric_data的结构
                    if isinstance(metric_data, dict):
                        # 如果包含daily_data键，这是Excel读取器的标准格式
                        if 'daily_data' in metric_data:
                            daily_data = metric_data['daily_data']
                            if isinstance(daily_data, dict):
                                logger.debug(f"指标 {metric_name} 包含日期: {list(daily_data.keys())}")
                                excel_dates.update(daily_data.keys())
                        # 否则，可能是直接的日期字典
                        else:
                            logger.debug(f"指标 {metric_name} 直接包含日期: {list(metric_data.keys())}")
                            excel_dates.update(metric_data.keys())
                    # 如果是字符串且看起来像字典，尝试解析
                    elif isinstance(metric_data, str) and metric_data.strip().startswith('{'):
                        try:
                            import ast
                            parsed_data = ast.literal_eval(metric_data)
                            if isinstance(parsed_data, dict):
                                logger.debug(f"从字符串解析的指标 {metric_name} 包含日期: {list(parsed_data.keys())}")
                                excel_dates.update(parsed_data.keys())
                        except (ValueError, SyntaxError) as e:
                            logger.debug(f"无法解析指标 {metric_name} 的字符串数据: {e}")
            
            # 安全地记录日期信息
            date_list = sorted(list(excel_dates))
            logger.info(f"从Excel中提取到的日期数量: {len(date_list)}")
            logger.info(f"日期详情: {[d.encode('utf-8').decode('utf-8') for d in date_list[:10]]}")  # 只显示前10个
            
            if not excel_dates:
                return {"has_coverage_issues": False, "message": "Excel中没有找到日期数据"}
            
            # 检查数据库文件是否存在
            from datetime import datetime, timedelta
            import os
            from services.database_manager import get_db_manager
            
            year, month = map(int, self.target_month.split('-'))
            start_date = datetime(year, month, 1)
            
            # 获取该月的最后一天
            if month == 12:
                end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(days=1)
            
            # 检查每个日期的数据库文件是否存在（考虑日期偏移）
            available_dates = []
            missing_dates = []
            
            logger.info(f"开始检查 {self.target_month} 月份的数据库文件（考虑日期偏移）")
            
            current_date = start_date
            while current_date <= end_date:
                try:
                    # 使用安全的日期格式化
                    day = current_date.day
                    date_str = f"{day}日"
                    
                    # 检查这个日期是否在Excel数据中
                    if date_str in excel_dates:
                        # 计算对应的数据库日期（Excel日期-1天）
                        db_date = current_date - timedelta(days=1)
                        
                        # 检查对应的数据库文件是否存在
                        db_path = get_db_manager().get_database_path(db_date)
                        logger.debug(f"检查Excel日期 {day}日 对应的数据库文件（{db_date.strftime('%Y-%m-%d')}）: {db_path}")
                        
                        if os.path.exists(db_path):
                            available_dates.append(date_str)
                            logger.debug("  ✓ 文件存在")
                        else:
                            missing_dates.append(date_str)
                            logger.debug("  ✗ 文件不存在")
                    else:
                        logger.debug(f"日期 {day}日 不在Excel数据中，跳过检查")
                        
                except Exception as date_error:
                    logger.warning(f"处理日期 {current_date} 时出错: {date_error}")
                
                current_date += timedelta(days=1)
            
            logger.info(f"检查完成 - 可用日期: {len(available_dates)}, 缺失日期: {len(missing_dates)}")
            
            # 生成覆盖报告
            excel_dates_count = len(excel_dates)
            available_dates_count = len(available_dates)
            missing_dates_count = len(missing_dates)
            
            # 计算覆盖率，避免除零错误
            coverage_rate = 0
            if excel_dates_count > 0:
                coverage_rate = available_dates_count / excel_dates_count
            
            coverage_info = {
                "target_month": self.target_month,
                "excel_dates": sorted(list(excel_dates)),
                "available_dates": sorted(available_dates),
                "missing_dates": sorted(missing_dates),
                "has_coverage_issues": missing_dates_count > 0 or excel_dates_count == 0,
                "coverage_rate": coverage_rate,
                "excel_dates_count": excel_dates_count,
                "available_dates_count": available_dates_count,
                "missing_dates_count": missing_dates_count
            }
            
            if missing_dates:
                coverage_info["message"] = f"数据库文件不完整：Excel中有 {len(excel_dates)} 个日期，数据库中缺少 {len(missing_dates)} 个日期的数据文件"
                coverage_info["suggestion"] = "建议先执行批量同步数据操作，确保所有日期的数据库文件都已生成"
                coverage_info["date_offset_note"] = "注意：Excel中的日期对应数据库中的前一天数据（如Excel的8月2日对应数据库的8月1日）"
            else:
                coverage_info["message"] = f"数据库文件完整：{len(available_dates)} 个日期的数据文件都已存在"
                coverage_info["date_offset_note"] = "注意：Excel中的日期对应数据库中的前一天数据（如Excel的8月2日对应数据库的8月1日）"
            
            logger.info(f"数据库覆盖检查完成: {coverage_info['message']}")
            
            return coverage_info
            
        except Exception as e:
            error_msg = str(e)
            logger.error(f"检查数据库覆盖异常: {error_msg}")
            return {
                "has_coverage_issues": True,
                "message": f"检查数据库覆盖时发生错误: {error_msg}",
                "coverage_rate": 0,
                "excel_dates_count": 0,
                "available_dates_count": 0,
                "missing_dates_count": 0,
                "suggestion": "请检查系统配置或联系技术支持"
            }
    
    async def _compare_store_data(self, excel_store_name: str, excel_data: Dict, db_store: Dict, selected_fields: List[str]) -> Dict:
        """
        对比单个门店的数据
        
        Args:
            excel_store_name: Excel门店名称
            excel_data: Excel门店数据
            db_store: 数据库门店信息
            selected_fields: 选中的字段列表
            
        Returns:
            Dict: 门店对比结果
        """
        try:
            store_comparison = {
                "excel_store_name": excel_store_name,
                "db_store_name": db_store['storeName'],
                "db_store_id": db_store.get('id'),
                "daily_comparisons": {},
                "has_data_errors": False,
                "data_errors": []
            }
            
            # 获取该门店的数据库数据
            db_data = await self._get_store_database_data(db_store['storeName'])
            
            if not db_data:
                self.warnings.append({
                    "type": "no_database_data",
                    "store_name": excel_store_name,
                    "message": f"门店 {excel_store_name} 在数据库中没有数据"
                })
                return store_comparison
            
            # 对比每日数据
            for metric_name, metric_data in excel_data.items():
                if not isinstance(metric_data, dict):
                    continue
                
                # 获取对应的字段键名
                field_key = self._get_field_key_by_excel_name(metric_name)
                if not field_key or field_key not in selected_fields:
                    continue
                
                # 提取每日数据
                daily_data = metric_data.get('daily_data', metric_data)
                if not isinstance(daily_data, dict):
                    continue
                
                # 对比该指标的每日数据
                for date_str, excel_value in daily_data.items():
                    # 调试：记录Excel中的日期格式
                    logger.debug(f"Excel日期: '{date_str}' (类型: {type(date_str)})")
                    logger.debug(f"数据库可用日期: {list(db_data.keys())}")
                    
                    # 数据库数据已经按Excel日期格式组织，直接使用Excel日期作为键
                    # 只处理数据库中有数据的日期
                    if date_str not in db_data:
                        logger.debug(f"跳过Excel日期 {date_str}，数据库中无对应数据")
                        continue
                    
                    if date_str not in store_comparison["daily_comparisons"]:
                        store_comparison["daily_comparisons"][date_str] = {}
                    
                    # 获取数据库中对应的值（直接使用Excel日期作为键）
                    db_value = self._get_db_value(db_data, date_str, field_key)
                    
                    # 记录对比结果（保持原始格式）
                    comparison = {
                        "excel_value": excel_value,
                        "db_value": db_value,
                        "is_different": self._values_are_different(excel_value, db_value)
                    }
                    
                    store_comparison["daily_comparisons"][date_str][field_key] = comparison
                    
                    # 如果数据不一致，记录错误
                    if comparison["is_different"]:
                        store_comparison["has_data_errors"] = True
                        store_comparison["data_errors"].append({
                            "date": date_str,
                            "field": field_key,
                            "excel_value": excel_value,
                            "db_value": db_value
                        })
            
            return store_comparison
            
        except Exception as e:
            logger.error(f"对比门店数据异常: {e}")
            return {}
    
    def _get_field_key_by_excel_name(self, excel_name: str) -> Optional[str]:
        """根据Excel指标名称获取字段键名"""
        for excel_metric_name, field_key in EXCEL_METRICS_MAPPING.items():
            if excel_metric_name == excel_name:
                return field_key
        return None
    
    async def _get_store_database_data(self, store_name: str) -> Dict:
        """获取门店的数据库数据"""
        try:
            if not self.db_manager:
                return {}
            
            # 查询该门店的所有数据
            # 使用模糊查询来匹配门店名称
            from datetime import datetime, timedelta
            import os
            
            # 如果有target_month，查询该月份及上个月的数据；否则查询当前月份
            if self.target_month:
                # 解析target_month (例如: "2025-09")
                year, month = map(int, self.target_month.split('-'))
                
                # 查询上个月最后一天到目标月份最后一天的数据
                # 因为Excel的1日可能对应上个月最后一天的数据
                if month == 1:
                    # 如果目标月份是1月，查询上一年12月最后一天到1月31日
                    start_date = datetime(year - 1, 12, 1)
                    # 获取上一年12月的最后一天
                    start_date = (start_date.replace(month=1, year=year) - timedelta(days=1))
                    end_date = datetime(year, 2, 1) - timedelta(days=1)
                else:
                    # 获取上个月的最后一天
                    start_date = datetime(year, month, 1) - timedelta(days=1)
                    end_date = datetime(year, month + 1, 1) - timedelta(days=1)
            else:
                start_date = datetime.now()
                end_date = start_date
            
            # 收集该月份所有天的数据
            all_data = []
            current_date = start_date
            
            while current_date <= end_date:
                result = self.db_manager.query_data(
                    date=current_date,
                    mdshow_filter=store_name,
                    page=1,
                    page_size=1000  # 获取足够多的数据
                )
                
                if result and result.get('data'):
                    all_data.extend(result['data'])
                
                current_date += timedelta(days=1)
            
            if not all_data:
                logger.warning(f"门店 {store_name} 在 {self.target_month} 月份没有找到数据")
                return {}
            
            # 将数据按日期组织
            organized_data = {}
            for record in all_data:
                create_date = record.get('createDateShow', '')
                if not create_date:
                    continue
                
                # 调试：记录数据库中的日期格式
                logger.debug(f"数据库记录日期: '{create_date}' (类型: {type(create_date)})")
                
                # 将数据库日期格式转换为Excel格式，考虑日期偏移
                # 数据库格式: '2025-07-31' -> Excel格式: '1日'（Excel日期对应数据库前一天）
                excel_date_key = self._convert_db_date_to_excel_format_with_offset(create_date)
                if not excel_date_key:
                    continue
                
                if excel_date_key not in organized_data:
                    organized_data[excel_date_key] = {}
                
                # 将记录的所有字段都保存
                for field_key, value in record.items():
                    if field_key in FIELD_CONFIG:
                        organized_data[excel_date_key][field_key] = value
            
            logger.info(f"门店 {store_name} 数据库数据组织完成，包含日期: {list(organized_data.keys())}")
            return organized_data
            
        except Exception as e:
            logger.error(f"获取门店数据库数据异常: {e}")
            return {}
    
    def _convert_db_date_to_excel_format(self, db_date: str) -> str:
        """将数据库日期格式转换为Excel格式，考虑日期偏移"""
        try:
            if not db_date:
                return ""
            
            # 数据库格式: '2025-08-01' -> Excel格式: '1日'（数据库日期直接对应Excel日期）
            if '-' in db_date:
                # 提取日期部分
                date_part = db_date.split(' ')[0]  # 处理可能的时间部分
                parts = date_part.split('-')
                if len(parts) >= 3:
                    from datetime import datetime, timedelta
                    # 解析数据库日期
                    db_datetime = datetime.strptime(date_part, '%Y-%m-%d')
                    # 数据库日期直接对应Excel日期（不需要偏移）
                    return f"{db_datetime.day}日"
            
            # 如果已经是Excel格式，直接返回
            if db_date.endswith('日'):
                return db_date
            
            return ""
            
        except Exception as e:
            logger.debug(f"日期格式转换异常: {e}")
            return ""
    
    def _convert_db_date_to_excel_format_with_offset(self, db_date: str) -> str:
        """将数据库日期格式转换为Excel格式，考虑日期偏移（Excel日期对应数据库前一天）"""
        try:
            if not db_date:
                return ""
            
            # 数据库格式: '2025-07-31' -> Excel格式: '1日'（Excel日期对应数据库前一天）
            if '-' in db_date:
                # 提取日期部分
                date_part = db_date.split(' ')[0]  # 处理可能的时间部分
                parts = date_part.split('-')
                if len(parts) >= 3:
                    from datetime import datetime, timedelta
                    # 解析数据库日期
                    db_datetime = datetime.strptime(date_part, '%Y-%m-%d')
                    # 加一天，因为Excel中的日期对应数据库前一天的数据
                    excel_datetime = db_datetime + timedelta(days=1)
                    return f"{excel_datetime.day}日"
            
            # 如果已经是Excel格式，直接返回
            if db_date.endswith('日'):
                return db_date
            
            return ""
            
        except Exception as e:
            logger.debug(f"日期格式转换异常: {e}")
            return ""
    
    def _convert_excel_date_to_db_key(self, excel_date: str) -> str:
        """将Excel日期格式转换为数据库日期键，考虑日期偏移"""
        try:
            if not excel_date or not self.target_month:
                return excel_date
            
            # 提取Excel日期中的天数
            if excel_date.endswith('日'):
                day = int(excel_date[:-1])
                
                # 解析目标月份
                year, month = map(int, self.target_month.split('-'))
                
                # 创建Excel日期对象
                from datetime import datetime, timedelta
                excel_datetime = datetime(year, month, day)
                
                # 减一天得到对应的数据库日期
                db_datetime = excel_datetime - timedelta(days=1)
                
                # 返回数据库日期格式
                return db_datetime.strftime('%Y-%m-%d')
            
            return excel_date
            
        except Exception as e:
            logger.debug(f"Excel日期转换为数据库键异常: {e}")
            return excel_date
    
    def _get_db_value(self, db_data: Dict, date_str: str, field_key: str) -> Any:
        """从数据库数据中获取指定日期和字段的值"""
        try:
            # 根据实际的数据库数据结构来实现
            if date_str in db_data and field_key in db_data[date_str]:
                return db_data[date_str][field_key]
            return None
            
        except Exception as e:
            logger.debug(f"获取数据库值异常: {e}")
            return None
    
    def _convert_to_percentage(self, value: Any) -> str:
        """将数值转换为百分比格式"""
        try:
            if value is None or value == "":
                return ""
            
            # 如果已经是百分比格式，直接返回
            if isinstance(value, str) and value.endswith('%'):
                return value
            
            # 转换为数字
            num_value = float(value)
            
            # 转换为百分比（乘以100并四舍五入到2位小数）
            percentage = round(num_value * 100, 2)
            return f"{percentage}%"
            
        except (ValueError, TypeError):
            return str(value)
    
    def _normalize_value_for_comparison(self, value: Any) -> float:
        """标准化值用于比较"""
        try:
            if value is None or value == "":
                return 0.0
            
            value_str = str(value).strip()
            
            # 处理百分比格式
            if value_str.endswith('%'):
                return float(value_str[:-1])
            
            # 处理小数格式（转换为百分比数值）
            num_value = float(value_str)
            if 0 <= num_value <= 1:
                # 小数格式，转换为百分比
                return num_value * 100
            else:
                # 已经是百分比数值或其他格式
                return num_value
            
        except (ValueError, TypeError):
            return 0.0
    
    def _values_are_different(self, excel_value: Any, db_value: Any) -> bool:
        """判断两个值是否不同，保持原始格式进行比较"""
        try:
            # 处理None值
            if excel_value is None and db_value is None:
                return False
            if excel_value is None or db_value is None:
                return True
            
            # 标准化值进行比较
            excel_normalized = self._normalize_value_for_comparison(excel_value)
            db_normalized = self._normalize_value_for_comparison(db_value)
            
            # 使用较小的精度进行比较（考虑到四舍五入误差）
            return abs(excel_normalized - db_normalized) > 0.0001
                
        except Exception as e:
            logger.debug(f"值比较异常: {e}")
            return True
    
    async def export_comparison_excel(self, comparison_data: Dict) -> Dict[str, Any]:
        """
        导出对比结果到Excel文件
        
        Args:
            comparison_data: 对比数据
            
        Returns:
            Dict: 导出结果
        """
        try:
            logger.info("开始生成对比Excel文件")
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "数据对比报告"
            
            # 设置样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            error_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            error_font = Font(color="FFFFFF", bold=True)
            
            # 生成表头
            self._create_comparison_headers(ws, comparison_data)
            
            # 填充数据
            self._fill_comparison_data(ws, comparison_data)
            
            # 应用样式
            self._apply_comparison_styles(ws, header_fill, header_font, error_fill, error_font)
            
            # 生成基于字段配置和目标月份的固定文件名
            from backend.constants.field_config import FIELD_CONFIG
            field_rule = config_db_manager.get_export_field_rule()
            if field_rule and field_rule.get('selected_fields'):
                selected_fields = field_rule['selected_fields']
            else:
                selected_fields = list(FIELD_CONFIG.keys())
            
            # 基于字段配置生成哈希值
            import hashlib
            fields_hash = hashlib.md5(''.join(sorted(selected_fields)).encode()).hexdigest()[:8]
            
            # 使用目标月份而不是当前日期
            if self.target_month:
                # 从 "2025-08" 格式转换为 "25_08" 格式
                year, month = self.target_month.split('-')
                date_part = f"{year[-2:]}_{month}"
            else:
                # 如果没有目标月份，使用当前月份
                date_part = datetime.now().strftime("%y_%m")
            
            filename = f"ksx_{date_part}_{fields_hash}.xlsx"
            
            # 获取导出目录
            export_dir = self._get_export_dir()
            os.makedirs(export_dir, exist_ok=True)
            
            file_path = os.path.join(export_dir, filename)
            wb.save(file_path)
            
            # 生成摘要
            summary = self._generate_export_summary(comparison_data)
            
            logger.info(f"对比Excel文件生成成功: {file_path}")
            
            return {
                "file_path": file_path,
                "file_name": filename,
                "summary": summary
            }
            
        except Exception as e:
            logger.error(f"导出对比Excel异常: {e}")
            raise e
    
    def _get_export_dir(self) -> str:
        """获取导出目录"""
        # 复用export.py中的get_app_data_dir逻辑
        if getattr(sys, 'frozen', False):
            # PyInstaller打包后的情况
            executable_path = sys.executable
            if "/Applications/" in executable_path:
                # DMG安装版本：导出到下载文件夹
                return os.path.expanduser("~/Downloads")
            else:
                # 开发版本：使用dist目录
                if "KSX门店管理系统" in executable_path:
                    # 找到dist目录
                    parts = executable_path.split(os.sep)
                    for i, part in enumerate(parts):
                        if part == "dist":
                            dist_dir = os.sep.join(parts[:i+1])
                            return os.path.join(dist_dir, "data")
                # 如果找不到dist目录，使用默认路径
                return os.path.join(os.path.dirname(sys.executable), "data")
        else:
            # 开发环境
            return os.path.join(project_root, "backend", "exports")
    
    def _sort_dates_by_number(self, dates: List[str]) -> List[str]:
        """按日期数字排序，而不是字符串排序"""
        def sort_key(date_str):
            """提取日期数字进行排序"""
            try:
                if date_str.endswith('日'):
                    day_num = int(date_str[:-1])
                    return day_num
                return 0
            except:
                return 0
        
        return sorted(dates, key=sort_key)
    
    def _create_comparison_headers(self, ws, comparison_data: Dict):
        """创建对比表格的表头"""
        try:
            # 获取所有日期
            all_dates = set()
            for store_data in comparison_data.values():
                all_dates.update(store_data.get("daily_comparisons", {}).keys())
            
            # 按日期数字排序，而不是字符串排序
            sorted_dates = self._sort_dates_by_number(list(all_dates))
            
            # 设置基础表头（合并单元格）
            ws.cell(row=1, column=1, value="序号")
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            
            ws.cell(row=1, column=2, value="门店名称")
            ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
            
            ws.cell(row=1, column=3, value="指标名称")
            ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
            
            # 为每个日期创建双列（系统数据 + 填报数据）
            current_col = 4
            for date in sorted_dates:
                # 第一行：日期（跨两列）
                ws.cell(row=1, column=current_col, value=date)
                ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 1)
                
                # 第二行：系统/填报标识
                ws.cell(row=2, column=current_col, value="系统")
                ws.cell(row=2, column=current_col + 1, value="填报")
                
                current_col += 2
            
            logger.info(f"创建表头完成，日期数量: {len(sorted_dates)}, 总列数: {current_col - 1}")
            
        except Exception as e:
            logger.error(f"创建表头异常: {e}")
            raise e
    
    def _fill_comparison_data(self, ws, comparison_data: Dict):
        """填充对比数据"""
        try:
            # 获取字段配置
            field_rule = config_db_manager.get_export_field_rule()
            if field_rule and field_rule.get('selected_fields'):
                selected_fields = field_rule['selected_fields']
            else:
                selected_fields = list(FIELD_CONFIG.keys())
            
            # 获取所有日期并排序
            all_dates = set()
            for store_data in comparison_data.values():
                all_dates.update(store_data.get("daily_comparisons", {}).keys())
            
            # 按日期数字排序，而不是字符串排序
            sorted_dates = self._sort_dates_by_number(list(all_dates))
            
            current_row = 3  # 从第3行开始填充数据
            serial_number = 1
            
            for store_name, store_data in comparison_data.items():
                store_start_row = current_row
                db_store_name = store_data.get("db_store_name", store_name)
                daily_comparisons = store_data.get("daily_comparisons", {})
                
                # 为每个选中的字段创建一行
                for field_key in selected_fields:
                    field_display_name = get_field_display_name(field_key)
                    
                    # 基础信息列
                    ws.cell(row=current_row, column=1, value=serial_number)
                    ws.cell(row=current_row, column=2, value=f"{store_name}\n({db_store_name})")
                    ws.cell(row=current_row, column=3, value=field_display_name)
                    
                    # 填充每日数据
                    current_col = 4
                    for date in sorted_dates:
                        daily_data = daily_comparisons.get(date, {})
                        field_comparison = daily_data.get(field_key, {})
                        
                        # 获取原始值（保持原始格式）
                        db_value = field_comparison.get("db_value", "")
                        excel_value = field_comparison.get("excel_value", "")
                        is_different = field_comparison.get("is_different", False)
                        
                        # 调试信息
                        logger.debug(f"填充数据 - 门店: {store_name}, 字段: {field_key}, 日期: {date}")
                        logger.debug(f"  db_value: {db_value} (类型: {type(db_value)})")
                        logger.debug(f"  excel_value: {excel_value} (类型: {type(excel_value)})")
                        
                        # 确保值是Excel兼容的类型，保持原始格式
                        db_display_value = self._convert_to_excel_value(db_value)
                        excel_display_value = self._convert_to_excel_value(excel_value)
                        
                        # 系统数据列
                        db_cell = ws.cell(row=current_row, column=current_col, value=db_display_value)
                        # 填报数据列
                        excel_cell = ws.cell(row=current_row, column=current_col + 1, value=excel_display_value)
                        
                        # 如果数据不一致，标红
                        if is_different:
                            db_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            db_cell.font = Font(color="FFFFFF", bold=True)
                            excel_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            excel_cell.font = Font(color="FFFFFF", bold=True)
                        
                        current_col += 2
                    
                    current_row += 1
                
                # 合并门店名称单元格
                if current_row > store_start_row + 1:
                    ws.merge_cells(start_row=store_start_row, start_column=2, end_row=current_row - 1, end_column=2)
                
                serial_number += 1
            
            logger.info(f"数据填充完成，总行数: {current_row - 1}")
            
        except Exception as e:
            logger.error(f"填充数据异常: {e}")
            raise e
    
    def _apply_comparison_styles(self, ws, header_fill, header_font, error_fill, error_font):
        """应用对比样式"""
        try:
            # 设置边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置对齐方式
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 应用表头样式
            max_row = ws.max_row
            max_col = ws.max_column
            
            # 第一行和第二行表头样式
            for row in range(1, 3):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = thin_border
            
            # 数据行样式
            for row in range(3, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':
                        # 只有没有填充色的单元格才添加边框和对齐
                        cell.border = thin_border
                        cell.alignment = center_alignment
            
            # 设置列宽
            ws.column_dimensions['A'].width = 8   # 序号
            ws.column_dimensions['B'].width = 25  # 门店名称
            ws.column_dimensions['C'].width = 20  # 指标名称
            
            # 设置日期列宽度
            for col in range(4, max_col + 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 12
            
            # 设置行高
            for row in range(1, max_row + 1):
                ws.row_dimensions[row].height = 25
            
            logger.info("样式应用完成")
            
        except Exception as e:
            logger.error(f"应用样式异常: {e}")
            raise e
    
    def _generate_export_summary(self, comparison_data: Dict) -> Dict:
        """生成导出摘要"""
        total_stores = len(comparison_data)
        stores_with_errors = sum(1 for store_data in comparison_data.values() if store_data.get("has_data_errors", False))
        
        return {
            "total_stores": total_stores,
            "stores_with_errors": stores_with_errors,
            "total_errors": len(self.errors),
            "total_warnings": len(self.warnings)
        }
    
    async def process_comparison(self, excel_data: Dict, stores: List[Dict], selected_fields: List[str] = None) -> Dict:
        """
        处理数据对比的主要方法
        
        Args:
            excel_data: Excel数据
            stores: 门店列表
            selected_fields: 选中的字段列表
            
        Returns:
            Dict: 对比结果
        """
        try:
            logger.info(f"开始处理数据对比，目标月份: {self.target_month}")
            
            # 重置状态
            self.errors = []
            self.warnings = []
            self.comparison_data = {}
            
            # 检查数据库覆盖率
            database_info = await self._check_database_coverage(excel_data)
            logger.info(f"数据库覆盖检查完成: {database_info.get('message', 'N/A')}")
            logger.info(f"数据库信息详情: {database_info}")
            
            # 如果没有选择字段，使用默认字段
            if not selected_fields:
                from backend.constants.field_config import FIELD_CONFIG
                selected_fields = list(FIELD_CONFIG.keys())
            
            # 对比每个门店的数据
            for store_name, store_data in excel_data.items():
                logger.info(f"处理Excel门店: {store_name}")
                
                # 匹配数据库门店
                matched_stores = await self._find_matching_stores(store_name)
                
                if not matched_stores:
                    self.errors.append({
                        "type": "store_not_found",
                        "excel_store": store_name,
                        "message": f"Excel门店 '{store_name}' 在数据库中未找到匹配"
                    })
                    continue
                elif len(matched_stores) > 1:
                    self.warnings.append({
                        "type": "multiple_matches",
                        "excel_store": store_name,
                        "matched_stores": [store['storeName'] for store in matched_stores],
                        "message": f"Excel门店 '{store_name}' 匹配到多个数据库门店"
                    })
                    # 使用第一个匹配的门店
                    db_store = matched_stores[0]
                else:
                    db_store = matched_stores[0]
                
                logger.info(f"Excel门店 '{store_name}' 匹配到数据库门店 '{db_store['storeName']}'")
                
                # 对比门店数据
                store_comparison = await self._compare_store_data(
                    store_name, store_data, db_store, selected_fields
                )
                
                if store_comparison:
                    self.comparison_data[store_name] = store_comparison
            
            # 检查是否有错误
            has_errors = len(self.errors) > 0 or len(self.warnings) > 0
            
            # 记录返回给前端的数据库信息
            logger.info(f"返回给前端的数据库信息: coverage_rate={database_info.get('coverage_rate', 'N/A')}, "
                       f"excel_dates_count={database_info.get('excel_dates_count', 'N/A')}, "
                       f"available_dates_count={database_info.get('available_dates_count', 'N/A')}")
            
            return {
                "has_errors": has_errors,
                "errors": self.errors,
                "warnings": self.warnings,
                "comparison_data": self.comparison_data,
                "database_info": database_info
            }
            
        except Exception as e:
            logger.error(f"处理数据对比异常: {e}")
            raise e
