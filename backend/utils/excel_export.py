#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel导出工具模块
"""

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import io
import re
import os
import sys
from datetime import datetime
from typing import List, Dict, Any, Optional
from loguru import logger

def get_excel_export_dir():
    """获取Excel导出目录，支持打包后的应用"""
    if getattr(sys, 'frozen', False):
        # PyInstaller打包后的情况
        app_dir = os.path.dirname(sys.executable)
        return os.path.join(app_dir, "data")
    else:
        # 开发环境
        current_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(current_dir, "exports")


def create_incremental_excel(monthly_data: dict, selected_fields: List[str], rule_name: str, filename: str) -> bytes:
    """创建支持增量更新的Excel文件"""
    
    # 使用正确的导出目录
    exports_dir = get_excel_export_dir()
    os.makedirs(exports_dir, exist_ok=True)
    existing_file_path = os.path.join(exports_dir, filename)
    wb = None
    
    if os.path.exists(existing_file_path):
        try:
            # 加载现有文件
            wb = load_workbook(existing_file_path)
            logger.info(f"加载现有Excel文件: {existing_file_path}")
        except Exception as e:
            logger.warning(f"无法加载现有文件 {existing_file_path}: {e}")
            wb = None
    
    if wb is None:
        # 创建新的工作簿
        wb = Workbook()
        # 删除默认工作表
        wb.remove(wb.active)
        logger.info(f"创建新的Excel文件: {filename}")
    
    # 处理每个月的数据
    for month, month_data in monthly_data.items():
        sheet_name = f"{month}月"
        
        # 检查工作表是否存在
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"更新现有工作表: {sheet_name}")
        else:
            ws = wb.create_sheet(title=sheet_name)
            logger.info(f"创建新工作表: {sheet_name}")
        
        # 更新工作表数据
        update_monthly_sheet(ws, month_data, selected_fields, month)
    
    # 确保导出目录存在
    os.makedirs(exports_dir, exist_ok=True)
    logger.info(f"创建导出目录: {exports_dir}")
    
    # 保存文件
    wb.save(existing_file_path)
    logger.info(f"Excel文件已保存: {existing_file_path}")
    
    # 检查文件是否真的存在
    if os.path.exists(existing_file_path):
        logger.info(f"文件确认存在: {existing_file_path}")
        # 返回文件内容
        with open(existing_file_path, 'rb') as f:
            return f.read()
    else:
        logger.error(f"文件保存失败: {existing_file_path}")
        raise Exception(f"文件保存失败: {existing_file_path}")


def update_monthly_sheet(ws, data: List[Dict], selected_fields: List[str], month: str):
    """更新月度工作表数据 - 增量添加列"""
    
    # 按门店和日期分组
    store_data = {}
    new_dates = set()
    
    for item in data:
        store_name = re.sub(r'<[^>]+>', '', item.get('MDShow', ''))
        create_date = item.get('createDateShow', '')
        
        if store_name and create_date:
            if store_name not in store_data:
                store_data[store_name] = {}
            store_data[store_name][create_date] = item
            new_dates.add(create_date)
    
    # 获取现有的列信息
    existing_columns = {}
    existing_stores = {}
    if ws.max_row > 0:
        # 读取第一行获取现有列
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and isinstance(cell_value, str):
                existing_columns[cell_value] = col
        
        # 读取第二列获取现有门店
        for row in range(2, ws.max_row + 1):
            store_name = ws.cell(row=row, column=2).value
            if store_name:
                existing_stores[store_name] = row
    
    # 检查是否有新的日期需要添加
    new_dates_to_add = []
    for date in new_dates:
        if date not in existing_columns:
            new_dates_to_add.append(date)
    
    # 如果没有新日期，直接返回
    if not new_dates_to_add:
        logger.info(f"所有日期都已存在，无需更新")
        return
    
    logger.info(f"需要添加的新日期: {new_dates_to_add}")
    
    # 获取所有现有日期和新日期，按顺序排列
    all_dates = sorted(list(existing_columns.keys()) + new_dates_to_add)
    
    # 获取所有门店（现有+新增）
    all_stores = sorted(set(list(existing_stores.keys()) + list(store_data.keys())))
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 检查是否需要设置基本表头
    need_header = False
    if ws.max_row == 0:
        need_header = True
    else:
        # 检查第1行第1列是否有"序号"
        first_cell = ws.cell(row=1, column=1).value
        if first_cell != "序号":
            need_header = True
    
    if need_header:
        # 第一列：序列号
        ws.cell(row=1, column=1, value="序号").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = header_alignment
        ws.cell(row=1, column=1).border = thin_border
        
        # 第二列：门店名称
        ws.cell(row=1, column=2, value="门店名称").font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        ws.cell(row=1, column=2).alignment = header_alignment
        ws.cell(row=1, column=2).border = thin_border
        
        # 第三列：指标名称
        ws.cell(row=1, column=3, value="指标名称").font = header_font
        ws.cell(row=1, column=3).fill = header_fill
        ws.cell(row=1, column=3).alignment = header_alignment
        ws.cell(row=1, column=3).border = thin_border
    
    # 添加新的日期列
    next_col = ws.max_column + 1 if ws.max_column > 0 else 3
    for date in sorted(new_dates_to_add):
        ws.cell(row=1, column=next_col, value=date).font = header_font
        ws.cell(row=1, column=next_col).fill = header_fill
        ws.cell(row=1, column=next_col).alignment = header_alignment
        ws.cell(row=1, column=next_col).border = thin_border
        existing_columns[date] = next_col
        next_col += 1
    
    # 填充数据
    if not existing_stores:
        # 如果是新文件，创建完整的表结构
        current_row = 2
        serial_number = 1
        for store_name in all_stores:
            store_start_row = current_row
            
            # 为每个字段创建行
            for field in selected_fields:
                field_name = FIELD_CONFIG.get(field, {}).get('name', field)
                
                # 设置序列号（合并单元格）
                ws.cell(row=store_start_row, column=1, value=serial_number)
                ws.cell(row=store_start_row, column=1).font = Font(bold=True)
                ws.cell(row=store_start_row, column=1).border = thin_border
                ws.cell(row=store_start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                
                # 设置门店名称（合并单元格）
                ws.cell(row=store_start_row, column=2, value=store_name)
                ws.cell(row=store_start_row, column=2).font = Font(bold=True)
                ws.cell(row=store_start_row, column=2).border = thin_border
                ws.cell(row=store_start_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
                
                # 设置指标名称
                ws.cell(row=store_start_row, column=3, value=field_name)
                ws.cell(row=store_start_row, column=3).border = thin_border
                ws.cell(row=store_start_row, column=3).alignment = Alignment(horizontal="left", vertical="center")
                
                # 填充所有日期的数据
                for date in all_dates:
                    if date in store_data.get(store_name, {}):
                        item = store_data[store_name][date]
                        value = item.get(field, '')
                        if field == 'id':
                            logger.info(f"ID字段处理: store_name={store_name}, date={date}, field={field}, value={value}, item_keys={list(item.keys())}")
                        col_index = existing_columns[date]
                        ws.cell(row=store_start_row, column=col_index, value=value)
                        ws.cell(row=store_start_row, column=col_index).border = thin_border
                        ws.cell(row=store_start_row, column=col_index).alignment = Alignment(horizontal="center", vertical="center")
                
                store_start_row += 1
            
            # 合并序列号和门店名称单元格
            merge_end_row = store_start_row - 1
            if merge_end_row > current_row:
                ws.merge_cells(f'A{current_row}:A{merge_end_row}')
                ws.merge_cells(f'B{current_row}:B{merge_end_row}')
            current_row = store_start_row
            serial_number += 1
    else:
        # 如果是现有文件，处理新日期和新门店
        # 1. 先处理现有门店的新日期数据
        for store_name in all_stores:
            store_row = existing_stores.get(store_name)
            if store_row is None:
                # 新门店，稍后处理
                continue
            
            # 为每个字段填充新日期的数据
            field_index = 0
            for field in selected_fields:
                row_index = store_row + field_index
                
                # 只填充新日期的数据
                for date in new_dates_to_add:
                    if date in store_data.get(store_name, {}):
                        item = store_data[store_name][date]
                        value = item.get(field, '')
                        col_index = existing_columns[date]
                        ws.cell(row=row_index, column=col_index, value=value)
                        ws.cell(row=row_index, column=col_index).border = thin_border
                        ws.cell(row=row_index, column=col_index).alignment = Alignment(horizontal="center", vertical="center")
                
                field_index += 1
        
        # 2. 处理新增门店（在文件底部添加）
        new_stores = [store for store in all_stores if store not in existing_stores]
        if new_stores:
            logger.info(f"发现新增门店: {new_stores}")
            current_row = ws.max_row + 1
            
            # 计算新增门店的起始序列号
            # 需要找到现有门店的最大序列号
            max_serial = 0
            for row in range(2, ws.max_row + 1):
                serial_cell = ws.cell(row=row, column=1).value
                if serial_cell and isinstance(serial_cell, (int, float)):
                    max_serial = max(max_serial, int(serial_cell))
            
            serial_number = max_serial + 1
            
            for store_name in new_stores:
                store_start_row = current_row
                
                # 为每个字段创建行
                for field in selected_fields:
                    field_name = FIELD_CONFIG.get(field, {}).get('name', field)
                    
                    # 设置序列号
                    ws.cell(row=store_start_row, column=1, value=serial_number)
                    ws.cell(row=store_start_row, column=1).font = Font(bold=True)
                    ws.cell(row=store_start_row, column=1).border = thin_border
                    ws.cell(row=store_start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                    
                    # 设置门店名称
                    ws.cell(row=store_start_row, column=2, value=store_name)
                    ws.cell(row=store_start_row, column=2).font = Font(bold=True)
                    ws.cell(row=store_start_row, column=2).border = thin_border
                    ws.cell(row=store_start_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
                    
                    # 设置指标名称
                    ws.cell(row=store_start_row, column=3, value=field_name)
                    ws.cell(row=store_start_row, column=3).border = thin_border
                    ws.cell(row=store_start_row, column=3).alignment = Alignment(horizontal="left", vertical="center")
                    
                    # 填充所有日期的数据
                    for date in all_dates:
                        if date in store_data.get(store_name, {}):
                            item = store_data[store_name][date]
                            value = item.get(field, '')
                            col_index = existing_columns[date]
                            ws.cell(row=store_start_row, column=col_index, value=value)
                            ws.cell(row=store_start_row, column=col_index).border = thin_border
                            ws.cell(row=store_start_row, column=col_index).alignment = Alignment(horizontal="center", vertical="center")
                    
                    store_start_row += 1
                
                # 合并序列号和门店名称单元格
                merge_end_row = store_start_row - 1
                if merge_end_row > current_row:
                    ws.merge_cells(f'A{current_row}:A{merge_end_row}')
                    ws.merge_cells(f'B{current_row}:B{merge_end_row}')
                current_row = store_start_row
                serial_number += 1
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8   # 序号
    ws.column_dimensions['B'].width = 25  # 门店名称
    ws.column_dimensions['C'].width = 20  # 指标名称
    for date in new_dates_to_add:
        col_letter = get_column_letter(existing_columns[date])
        ws.column_dimensions[col_letter].width = 12  # 日期列


# 字段配置和中文名称映射
FIELD_CONFIG = {
    "area": {"name": "区域", "comment": "门店所属区域"},
    "createDateShow": {"name": "创建日期", "comment": "数据创建日期"},
    "MDShow": {"name": "门店名称", "comment": "门店显示名称"},
    "totalScore": {"name": "总分", "comment": "门店综合评分"},
    "monthlyCanceledRate": {"name": "月度取消率", "comment": "月度订单取消率"},
    "dailyCanceledRate": {"name": "日取消率", "comment": "日订单取消率"},
    "monthlyMerchantRefundRate": {"name": "月度商家退款率", "comment": "月度商家责任退款率"},
    "monthlyOosRefundRate": {"name": "月度缺货退款率", "comment": "月度缺货退款率"},
    "monthlyJdOosRate": {"name": "月度京东缺货率", "comment": "月度京东缺货率"},
    "monthlyBadReviews": {"name": "月度差评数", "comment": "月度差评数量"},
    "monthlyBadReviewRate": {"name": "月度差评率", "comment": "月度差评率"},
    "monthlyPartialRefundRate": {"name": "月度部分退款率", "comment": "月度部分退款率"},
    "dailyMeituanRating": {"name": "美团评分", "comment": "美团平台日评分"},
    "dailyElemeRating": {"name": "饿了么评分", "comment": "饿了么平台日评分"},
    "dailyMeituanReplyRate": {"name": "美团回复率", "comment": "美团平台回复率"},
    "effectReply": {"name": "有效回复", "comment": "有效回复状态"},
    "monthlyMeituanPunctualityRate": {"name": "美团准时率", "comment": "美团月度准时送达率"},
    "monthlyElemeOntimeRate": {"name": "饿了么准时率", "comment": "饿了么月度准时送达率"},
    "monthlyJdFulfillmentRate": {"name": "京东履约率", "comment": "京东月度履约率"},
    "meituanComprehensiveExperienceDivision": {"name": "美团综合体验分", "comment": "美团综合体验评分"},
    "monthlyAvgStockRate": {"name": "月度平均库存率", "comment": "月度平均库存率"},
    "monthlyAvgTop500StockRate": {"name": "月度TOP500库存率", "comment": "月度TOP500商品库存率"},
    "monthlyAvgDirectStockRate": {"name": "月度直营库存率", "comment": "月度直营商品库存率"},
    "dailyTop500StockRate": {"name": "日TOP500库存率", "comment": "日TOP500商品库存率"},
    "dailyWarehouseSoldOut": {"name": "日仓库售罄数", "comment": "日仓库售罄商品数"},
    "dailyWarehouseStockRate": {"name": "日仓库库存率", "comment": "日仓库库存率"},
    "dailyDirectSoldOut": {"name": "日直营售罄数", "comment": "日直营售罄商品数"},
    "dailyDirectStockRate": {"name": "日直营库存率", "comment": "日直营库存率"},
    "dailyHybridSoldOut": {"name": "日混合售罄数", "comment": "日混合售罄商品数"},
    "dailyStockAvailability": {"name": "日库存可用率", "comment": "日库存可用率"},
    "dailyHybridStockRate": {"name": "日混合库存率", "comment": "日混合库存率"},
    "stockNoLocation": {"name": "无位置库存数", "comment": "无位置库存商品数"},
    "expiryManagement": {"name": "保质期管理", "comment": "保质期管理状态"},
    "inventoryLockOrders": {"name": "库存锁定订单", "comment": "库存锁定订单数"},
    "trainingCompleted": {"name": "培训完成", "comment": "培训完成状态"},
    "monthlyManhourPer100Orders": {"name": "月度百单工时", "comment": "月度每百单工时"},
    "monthlyTotalLoss": {"name": "月度总损失", "comment": "月度总损失金额"},
    "monthlyTotalLossRate": {"name": "月度总损失率", "comment": "月度总损失率"},
    "monthlyAvgDeliveryFee": {"name": "月度平均配送费", "comment": "月度平均配送费"},
    "dailyAvgDeliveryFee": {"name": "日平均配送费", "comment": "日平均配送费"},
    "monthlyCumulativeCancelRateScore": {"name": "月度累计取消率得分", "comment": "月度累计取消率得分"},
    "monthlyMerchantLiabilityRefundRateScore": {"name": "月度商家责任退款率得分", "comment": "月度商家责任退款率得分"},
    "monthlyStockoutRefundRateScore": {"name": "月度缺货退款率得分", "comment": "月度缺货退款率得分"},
    "monthlyNegativeReviewRateScore": {"name": "月度差评率得分", "comment": "月度差评率得分"},
    "monthlyPartialRefundRateScore": {"name": "月度部分退款率得分", "comment": "月度部分退款率得分"},
    "dailyMeituanRatingScore": {"name": "美团评分得分", "comment": "美团评分得分"},
    "dailyElemeRatingScore": {"name": "饿了么评分得分", "comment": "饿了么评分得分"},
    "monthlyMeituanDeliveryPunctualityRateScore": {"name": "美团配送准时率得分", "comment": "美团配送准时率得分"},
    "monthlyElemeTimelyDeliveryRateScore": {"name": "饿了么及时配送率得分", "comment": "饿了么及时配送率得分"},
    "validReplyWeightingPenalty": {"name": "有效回复权重惩罚", "comment": "有效回复权重惩罚"},
    "monthlyAverageStockRateWeightingPenalty": {"name": "月度平均库存率权重惩罚", "comment": "月度平均库存率权重惩罚"},
    "monthlyAverageTop500StockRateWeightingPenalty": {"name": "月度TOP500库存率权重惩罚", "comment": "月度TOP500库存率权重惩罚"},
    "monthlyAverageDirectStockRateWeightingPenalty": {"name": "月度直营库存率权重惩罚", "comment": "月度直营库存率权重惩罚"},
    "newProductComplianceListingWeightingPenalty": {"name": "新品合规上架权重惩罚", "comment": "新品合规上架权重惩罚"},
    "expiryManagementWeightingPenalty": {"name": "保质期管理权重惩罚", "comment": "保质期管理权重惩罚"},
    "inventoryLockWeightingPenalty": {"name": "库存锁定权重惩罚", "comment": "库存锁定权重惩罚"},
    "monthlyCumulativeHundredOrdersManhourWeightingPenalty": {"name": "月度累计百单工时权重惩罚", "comment": "月度累计百单工时权重惩罚"},
    "totalScoreWithoutWeightingPenalty": {"name": "无权重惩罚总分", "comment": "无权重惩罚总分"},
    "monthlyCumulativeMerchantLiabilityRefundRateWeightingPenalty": {"name": "月度累计商家责任退款率权重惩罚", "comment": "月度累计商家责任退款率权重惩罚"},
    "monthlyCumulativeOutOfStockRefundRateWeightingPenalty": {"name": "月度累计缺货退款率权重惩罚", "comment": "月度累计缺货退款率权重惩罚"},
    "meituanComplexExperienceScoreWeightingPenalty": {"name": "美团综合体验分权重惩罚", "comment": "美团综合体验分权重惩罚"},
    "meituanRatingWeightingPenalty": {"name": "美团评分权重惩罚", "comment": "美团评分权重惩罚"},
    "elemeRatingWeightingPenalty": {"name": "饿了么评分权重惩罚", "comment": "饿了么评分权重惩罚"},
    "partialRefundWeightingPenalty": {"name": "部分退款权重惩罚", "comment": "部分退款权重惩罚"},
    "trainingCompletedWeightingPenalty": {"name": "培训完成权重惩罚", "comment": "培训完成权重惩罚"},
    "totalWeightingPenalty": {"name": "总权重惩罚", "comment": "总权重惩罚"}
}


def create_excel_file(data: List[Dict[str, Any]], filename: str) -> str:
    """
    创建Excel文件
    
    Args:
        data: 要导出的数据列表
        filename: 文件名
    
    Returns:
        str: 文件路径
    """
    try:
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "KSX数据"
        
        if not data:
            logger.warning("没有数据可导出")
            return ""
        
        # 获取列名
        columns = list(data[0].keys())
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, column in enumerate(columns, 1):
                value = row_data.get(column, "")
                # 处理特殊字符
                if isinstance(value, str):
                    value = re.sub(r'[^\x00-\x7F]+', '', value)  # 移除非ASCII字符
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 设置列宽
        for col_idx, column in enumerate(columns, 1):
            max_length = len(str(column))
            for row_data in data:
                cell_value = str(row_data.get(column, ""))
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(data) + 1, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.border = thin_border
        
        # 保存文件
        file_path = f"backend/exports/{filename}"
        wb.save(file_path)
        logger.info(f"Excel文件已保存: {file_path}")
        
        return file_path
        
    except Exception as e:
        logger.error(f"创建Excel文件失败: {e}")
        return ""


def create_csv_content(data: List[Dict[str, Any]]) -> str:
    """
    创建CSV内容
    
    Args:
        data: 要导出的数据列表
    
    Returns:
        str: CSV内容
    """
    try:
        if not data:
            return ""
        
        # 使用pandas创建CSV
        df = pd.DataFrame(data)
        
        # 处理列名，移除特殊字符
        df.columns = [re.sub(r'[^\x00-\x7F]+', '', str(col)) for col in df.columns]
        
        # 转换为CSV字符串
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False, encoding='utf-8')
        csv_content = csv_buffer.getvalue()
        csv_buffer.close()
        
        return csv_content
        
    except Exception as e:
        logger.error(f"创建CSV内容失败: {e}")
        return ""


def generate_filename(prefix: str = "ksx_export", extension: str = "xlsx") -> str:
    """
    生成文件名
    
    Args:
        prefix: 文件名前缀
        extension: 文件扩展名
    
    Returns:
        str: 生成的文件名
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.{extension}"


