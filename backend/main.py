#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KSX数据查询API服务
基于FastAPI的后端服务，提供数据查询接口
"""

from fastapi import FastAPI, HTTPException, Query, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from datetime import datetime, date
import sys
import os
import asyncio
import subprocess
from loguru import logger
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import io
import re

# 添加项目根目录到路径，以便导入services模块
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(project_root)
from services.database_manager import get_db_manager
from services.config_database_manager import config_db_manager

# 创建FastAPI应用
app = FastAPI(
    title="KSX数据查询API",
    description="KSX门店数据查询服务",
    version="1.0.0"
)

# 配置CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 生产环境应该设置具体的域名
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 配置日志
logger.remove()
logger.add(
    sys.stderr,
    format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> | <level>{level: <8}</level> | <cyan>{name}</cyan> - <level>{message}</level>",
    level="INFO"
)
logger.add(
    "logs/api_{time:YYYY-MM-DD}.log",
    format="{time:YYYY-MM-DD HH:mm:ss} | {level: <8} | {name} - {message}",
    level="DEBUG",
    rotation="1 day",
    retention="30 days",
    compression="zip",
    encoding="utf-8"
)

# 请求模型
class SyncRequest(BaseModel):
    """同步请求模型"""
    date: Optional[str] = None  # 可选，如果不提供则使用默认日期

# 响应模型
class PageResponse(BaseModel):
    """分页响应模型"""
    data: List[Dict[str, Any]]
    total: int
    page: int
    page_size: int
    total_pages: int
    success: bool = True
    message: str = "查询成功"

class ErrorResponse(BaseModel):
    """错误响应模型"""
    success: bool = False
    message: str
    error_code: Optional[str] = None

class SyncResponse(BaseModel):
    """同步响应模型"""
    success: bool
    message: str
    total: Optional[int] = None
    error_code: Optional[str] = None

# API路由
@app.get("/", response_model=Dict[str, str])
async def root():
    """根路径"""
    return {
        "message": "KSX数据查询API服务",
        "version": "1.0.0",
        "docs_url": "/docs"
    }

@app.get("/api/health", response_model=Dict[str, str])
async def health_check():
    """健康检查"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat()
    }

@app.get("/api/data", response_model=PageResponse)
async def get_data(
    date_str: Optional[str] = Query(None, description="查询日期 (YYYY-MM-DD)，默认为今天"),
    mdshow: Optional[str] = Query(None, description="门店名称模糊查询"),
    page: int = Query(1, ge=1, description="页码，从1开始"),
    page_size: int = Query(20, ge=1, le=100, description="每页记录数，最大100")
):
    """
    获取KSX数据
    
    - **date_str**: 查询日期，格式YYYY-MM-DD，默认为今天
    - **mdshow**: 门店名称模糊查询，可选
    - **page**: 页码，从1开始
    - **page_size**: 每页记录数，最大100
    """
    try:
        # 调试：记录接收到的参数
        logger.info(f"接收到的参数: date_str={date_str}, mdshow={mdshow}, page={page}, page_size={page_size}")
        
        # 解析日期
        if date_str:
            try:
                query_date = datetime.strptime(date_str, "%Y-%m-%d")
                logger.info(f"解析日期参数: {date_str} -> {query_date}")
            except ValueError:
                raise HTTPException(
                    status_code=400, 
                    detail="日期格式错误，请使用YYYY-MM-DD格式"
                )
        else:
            query_date = datetime.now()
            logger.info(f"使用默认日期: {query_date}")
        
        # 获取数据库管理器
        db_manager = get_db_manager()
        
        # 调试：检查数据库路径
        db_path = db_manager.get_database_path(query_date)
        logger.info(f"查询数据库路径: {db_path}")
        logger.info(f"数据库文件是否存在: {db_path.exists()}")
        
        # 查询数据
        result = db_manager.query_data(
            date=query_date,
            mdshow_filter=mdshow,
            page=page,
            page_size=page_size
        )
        
        logger.info(f"查询数据: 日期={query_date.date()}, 门店={mdshow}, 页码={page}, 返回{len(result['data'])}条记录")
        
        # 调试：检查返回数据的日期
        if result['data']:
            first_record_date = result['data'][0].get('createDateShow', 'N/A')
            logger.info(f"返回的第一条记录日期: {first_record_date}")
        
        return PageResponse(
            data=result['data'],
            total=result['total'],
            page=result['page'],
            page_size=result['page_size'],
            total_pages=result['total_pages']
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"查询数据失败: {e}")
        raise HTTPException(status_code=500, detail=f"服务器内部错误: {str(e)}")

@app.get("/api/database/info", response_model=Dict[str, Any])
async def get_database_info():
    """获取数据库信息"""
    try:
        db_manager = get_db_manager()
        info = db_manager.get_database_info()
        
        logger.info("获取数据库信息")
        return {
            "success": True,
            "data": info
        }
        
    except Exception as e:
        logger.error(f"获取数据库信息失败: {e}")
        raise HTTPException(status_code=500, detail=f"服务器内部错误: {str(e)}")

@app.get("/api/data/search", response_model=PageResponse)
async def search_data(
    q: str = Query(..., description="搜索关键词（门店名称）"),
    date_str: Optional[str] = Query(None, description="查询日期 (YYYY-MM-DD)，默认为今天"),
    page: int = Query(1, ge=1, description="页码，从1开始"),
    page_size: int = Query(20, ge=1, le=100, description="每页记录数，最大100")
):
    """
    搜索门店数据
    
    这是一个便捷的搜索接口，等同于在get_data接口中设置mdshow参数
    """
    return await get_data(
        date_str=date_str,
        mdshow=q,
        page=page,
        page_size=page_size
    )

@app.get("/api/dates", response_model=Dict[str, Any])
async def get_available_dates():
    """获取有数据的日期列表"""
    try:
        db_manager = get_db_manager()
        info = db_manager.get_database_info()
        
        dates = []
        for month_info in info['months']:
            for db_info in month_info['databases']:
                # 从数据库文件名提取日期
                if db_info['name'].startswith('ksx_') and db_info['name'].endswith('.db'):
                    date_str = db_info['name'][4:-3]  # 移除 'ksx_' 前缀和 '.db' 后缀
                    dates.append(date_str)
        
        dates.sort(reverse=True)  # 按日期倒序排列
        
        logger.info(f"获取可用日期列表: {len(dates)}个日期")
        return {
            "success": True,
            "dates": dates,
            "total": len(dates)
        }
        
    except Exception as e:
        logger.error(f"获取日期列表失败: {e}")
        raise HTTPException(status_code=500, detail=f"服务器内部错误: {str(e)}")

async def run_crawler(target_date: Optional[str] = None):
    """运行爬虫程序"""
    try:
        if target_date:
            logger.info(f"开始运行爬虫程序，目标日期: {target_date}")
        else:
            logger.info("开始运行爬虫程序，使用默认日期")
        
        # 构建爬虫命令
        crawler_script = os.path.join(project_root, "services", "crawler", "main.py")
        
        # 使用uv运行爬虫
        cmd = ["uv", "run", "python", crawler_script]
        
        # 如果指定了日期，添加日期参数
        if target_date:
            cmd.extend(["--date", target_date])
        
        logger.info(f"执行命令: {' '.join(cmd)}")
        
        # 运行爬虫程序
        process = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            cwd=project_root
        )
        
        # 等待进程完成
        stdout, stderr = await process.communicate()
        
        if process.returncode == 0:
            output = stdout.decode('utf-8', errors='ignore')
            logger.info("爬虫程序执行成功")
            logger.info(f"输出: {output}")
            
            # 解析输出，获取最终数据条数和消息
            new_count = 0
            message = "数据同步完成"
            
            # 从输出中提取信息
            if "没有业务数据" in output or "当前日期没有业务数据" in output:
                message = "当前日期没有业务数据，请核查日期"
                new_count = 0
            elif "ERROR_TYPE: BROWSER_START_FAILED" in output:
                message = "浏览器启动失败，请检查浏览器安装或权限设置"
                new_count = 0
            elif "ERROR_TYPE: LOGIN_FAILED" in output:
                message = "登录失败，请检查用户名密码或网络连接"
                new_count = 0
            elif "ERROR_TYPE: DATA_EXTRACTION_FAILED" in output:
                message = "数据提取失败，请检查网络连接或稍后重试"
                new_count = 0
            elif "ERROR_TYPE: USER_INTERRUPTED" in output:
                message = "用户中断操作"
                new_count = 0
            elif "ERROR_TYPE: UNKNOWN_ERROR" in output:
                message = "未知错误，请查看日志或联系管理员"
                new_count = 0
            elif "共获取" in output:
                # 尝试从输出中提取数据条数
                import re
                match = re.search(r'共获取\s*(\d+)\s*条', output)
                if match:
                    new_count = int(match.group(1))
            elif "数据库记录" in output:
                # 尝试从输出中提取数据库记录数
                import re
                match = re.search(r'数据库记录:\s*(\d+)条', output)
                if match:
                    new_count = int(match.group(1))
            
            return True, message, new_count
        else:
            error_msg = stderr.decode('utf-8', errors='ignore')
            logger.error(f"爬虫程序执行失败: {error_msg}")
            return False, f"爬虫执行失败: {error_msg}", 0
            
    except Exception as e:
        logger.error(f"运行爬虫程序异常: {e}")
        return False, f"运行爬虫程序异常: {str(e)}", 0

@app.post("/api/sync-data", response_model=SyncResponse)
async def sync_data(request: SyncRequest, background_tasks: BackgroundTasks):
    """
    同步数据接口
    
    启动爬虫程序获取指定日期的数据
    """
    try:
        target_date = request.date
        if target_date:
            logger.info(f"收到同步数据请求，目标日期: {target_date}")
        else:
            logger.info("收到同步数据请求，使用默认日期")
        
        # 运行爬虫程序
        success, message, new_count = await run_crawler(target_date)
        
        if success:
            logger.info(f"数据同步完成，共 {new_count or 0} 条数据")
            return SyncResponse(
                success=True,
                message=message,
                total=new_count
            )
        else:
            logger.error(f"数据同步失败: {message}")
            return SyncResponse(
                success=False,
                message=message,
                error_code="CRAWLER_ERROR"
            )
        
    except Exception as e:
        logger.error(f"同步数据接口异常: {e}")
        return SyncResponse(
            success=False,
            message=f"同步数据失败: {str(e)}",
            error_code="INTERNAL_ERROR"
        )

# 门店配置相关接口
@app.get("/api/stores")
async def get_stores():
    """获取所有门店列表"""
    try:
        logger.info("开始获取门店列表")
        # 首先从配置数据库获取已存储的门店
        raw_stores = config_db_manager.get_all_stores()
        logger.info(f"原始门店数据: {raw_stores[:3] if raw_stores else '无数据'}")
        
        # 如果配置数据库中没有门店，从业务数据中提取门店列表
        if not raw_stores:
            db_manager = get_db_manager()
            
            # 获取最近30天的数据来提取门店
            from datetime import datetime, timedelta
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
            
            all_stores = set()
            current_date = start_date
            while current_date <= end_date:
                try:
                    result = db_manager.query_data(date=current_date, page=1, page_size=1000)
                    for item in result.get('data', []):
                        mdshow = item.get('MDShow', '')
                        if mdshow:
                            # 清理HTML标签
                            import re
                            clean_name = re.sub(r'<[^>]+>', '', mdshow)
                            if clean_name.strip():
                                all_stores.add(clean_name.strip())
                except Exception as e:
                    logger.warning(f"查询日期 {current_date.date()} 的数据失败: {e}")
                
                current_date += timedelta(days=1)
            
            # 将提取的门店添加到配置数据库
            for store_name in sorted(all_stores):
                config_db_manager.add_store(store_name)
            
            # 重新获取门店列表
            raw_stores = config_db_manager.get_all_stores()
        
        # 转换为前端期望的格式
        stores = []
        for i, store in enumerate(raw_stores):
            stores.append({
                'id': i + 1,  # 生成唯一ID
                'store_name': store['store_name'],  # 使用name字段作为store_name
                'created_at': '2025-01-01T00:00:00Z',  # 默认时间
                'updated_at': '2025-01-01T00:00:00Z'   # 默认时间
            })
        
        logger.info(f"转换后的门店数据: {stores[:3] if stores else '无数据'}")
        
        return {
            "success": True,
            "data": stores
        }
    except Exception as e:
        logger.error(f"获取门店列表失败: {e}")
        return {
            "success": False,
            "message": f"获取门店列表失败: {str(e)}"
        }

@app.get("/api/export-rule")
async def get_export_rule():
    """获取导出规则"""
    try:
        rule = config_db_manager.get_export_rule()
        return {
            "success": True,
            "data": rule
        }
    except Exception as e:
        logger.error(f"获取导出规则失败: {e}")
        return {
            "success": False,
            "message": f"获取导出规则失败: {str(e)}"
        }

@app.post("/api/export-rule")
async def save_export_rule(rule_data: dict):
    """保存导出规则"""
    try:
        selected_stores = rule_data.get("selected_stores", [])
        
        success = config_db_manager.save_export_rule(selected_stores)
        
        if success:
            return {
                "success": True,
                "message": "导出规则保存成功"
            }
        else:
            return {
                "success": False,
                "message": "导出规则保存失败"
            }
    except Exception as e:
        logger.error(f"保存导出规则失败: {e}")
        return {
            "success": False,
            "message": f"保存导出规则失败: {str(e)}"
        }

@app.post("/api/export-data")
async def export_data(export_config: dict = {}):
    """导出数据到CSV文件"""
    try:
        # 获取导出规则
        rule = config_db_manager.get_export_rule()
        if rule:
            selected_stores = rule.get('selected_stores', [])
            rule_name = rule.get('rule_name', '默认导出规则')
        else:
            selected_stores = []
            rule_name = '全部数据'
        
        # 获取数据库管理器
        db_manager = get_db_manager()
        
        # 查询数据
        if selected_stores:
            # 根据选中的门店查询数据
            store_names = []
            all_stores = config_db_manager.get_all_stores()
            store_id_to_name = {store['id']: store['store_name'] for store in all_stores}
            
            for store_id in selected_stores:
                if store_id in store_id_to_name:
                    store_names.append(store_id_to_name[store_id])
            
            if not store_names:
                return {
                    "success": False,
                    "message": "未找到有效的门店名称"
                }
            
            # 使用现有的query_data方法，但需要修改为支持多个门店
            # 暂时导出所有数据，然后在前端过滤
            result = db_manager.query_data(page=1, page_size=10000)  # 获取大量数据
            all_data = result.get('data', [])
            
            # 过滤选中的门店
            data = []
            import re
            for item in all_data:
                mdshow = item.get('MDShow', '')
                # 清理HTML标签
                clean_mdshow = re.sub(r'<[^>]+>', '', mdshow)
                if clean_mdshow in store_names:
                    data.append(item)
        else:
            # 导出所有数据
            result = db_manager.query_data(page=1, page_size=10000)  # 获取大量数据
            data = result.get('data', [])
        
        if not data:
            return {
                "success": False,
                "message": "没有找到数据"
            }
        
        # 生成CSV内容
        import csv
        from datetime import datetime
        import io
        
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ksx_export_{rule_name}_{timestamp}.csv"
        except Exception as e:
            logger.error(f"生成文件名失败: {e}, rule_name: {rule_name}, type: {type(rule_name)}")
            raise
        
        # 生成CSV内容到内存
        csv_buffer = io.StringIO()
        if data:
            fieldnames = data[0].keys()
            writer = csv.DictWriter(csv_buffer, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)
        
        csv_content = csv_buffer.getvalue()
        csv_buffer.close()
        
        logger.info(f"数据导出成功: {filename}, 共 {len(data)} 条记录")
        
        return {
            "success": True,
            "message": f"数据导出成功，共 {len(data)} 条记录",
            "filename": filename,
            "csv_content": csv_content,
            "count": len(data)
        }
        
    except Exception as e:
        logger.error(f"导出数据失败: {e}")
        return {
            "success": False,
            "message": f"导出数据失败: {str(e)}"
        }

# 字段配置和中文名称映射
FIELD_CONFIG = {
    "area": {"name": "区域", "comment": "门店所属区域"},
    "createDateShow": {"name": "创建日期", "comment": "数据创建日期"},
    "MDShow": {"name": "门店名称", "comment": "门店显示名称"},
    "totalScore": {"name": "总分", "comment": "门店综合评分"},
    "monthlyCanceledRate": {"name": "月度取消率", "comment": "月度订单取消率"},
    "dailyCanceledRate": {"name": "日取消率", "comment": "日订单取消率"},
    "monthlyMerchantRefundRate": {"name": "月度商家退款率", "comment": "月度商家责任退款率"},
    "monthlyOosRefundRate": {"name": "月度缺货退款率", "comment": "月度缺货退款率"},
    "monthlyJdOosRate": {"name": "月度京东缺货率", "comment": "月度京东缺货率"},
    "monthlyBadReviews": {"name": "月度差评数", "comment": "月度差评数量"},
    "monthlyBadReviewRate": {"name": "月度差评率", "comment": "月度差评率"},
    "monthlyPartialRefundRate": {"name": "月度部分退款率", "comment": "月度部分退款率"},
    "dailyMeituanRating": {"name": "美团评分", "comment": "美团平台日评分"},
    "dailyElemeRating": {"name": "饿了么评分", "comment": "饿了么平台日评分"},
    "dailyMeituanReplyRate": {"name": "美团回复率", "comment": "美团平台回复率"},
    "effectReply": {"name": "有效回复", "comment": "有效回复状态"},
    "monthlyMeituanPunctualityRate": {"name": "美团准时率", "comment": "美团月度准时送达率"},
    "monthlyElemeOntimeRate": {"name": "饿了么准时率", "comment": "饿了么月度准时送达率"},
    "monthlyJdFulfillmentRate": {"name": "京东履约率", "comment": "京东月度履约率"},
    "meituanComprehensiveExperienceDivision": {"name": "美团综合体验分", "comment": "美团综合体验评分"},
    "monthlyAvgStockRate": {"name": "月度平均库存率", "comment": "月度平均库存率"},
    "monthlyAvgTop500StockRate": {"name": "月度TOP500库存率", "comment": "月度TOP500商品库存率"},
    "monthlyAvgDirectStockRate": {"name": "月度直营库存率", "comment": "月度直营商品库存率"},
    "dailyTop500StockRate": {"name": "日TOP500库存率", "comment": "日TOP500商品库存率"},
    "dailyWarehouseSoldOut": {"name": "日仓库售罄数", "comment": "日仓库售罄商品数"},
    "dailyWarehouseStockRate": {"name": "日仓库库存率", "comment": "日仓库库存率"},
    "dailyDirectSoldOut": {"name": "日直营售罄数", "comment": "日直营售罄商品数"},
    "dailyDirectStockRate": {"name": "日直营库存率", "comment": "日直营库存率"},
    "dailyHybridSoldOut": {"name": "日混合售罄数", "comment": "日混合售罄商品数"},
    "dailyStockAvailability": {"name": "日库存可用率", "comment": "日库存可用率"},
    "dailyHybridStockRate": {"name": "日混合库存率", "comment": "日混合库存率"},
    "stockNoLocation": {"name": "无位置库存数", "comment": "无位置库存商品数"},
    "expiryManagement": {"name": "保质期管理", "comment": "保质期管理状态"},
    "inventoryLockOrders": {"name": "库存锁定订单", "comment": "库存锁定订单数"},
    "trainingCompleted": {"name": "培训完成", "comment": "培训完成状态"},
    "monthlyManhourPer100Orders": {"name": "月度百单工时", "comment": "月度每百单工时"},
    "monthlyTotalLoss": {"name": "月度总损失", "comment": "月度总损失金额"},
    "monthlyTotalLossRate": {"name": "月度总损失率", "comment": "月度总损失率"},
    "monthlyAvgDeliveryFee": {"name": "月度平均配送费", "comment": "月度平均配送费"},
    "dailyAvgDeliveryFee": {"name": "日平均配送费", "comment": "日平均配送费"},
    "monthlyCumulativeCancelRateScore": {"name": "月度累计取消率得分", "comment": "月度累计取消率得分"},
    "monthlyMerchantLiabilityRefundRateScore": {"name": "月度商家责任退款率得分", "comment": "月度商家责任退款率得分"},
    "monthlyStockoutRefundRateScore": {"name": "月度缺货退款率得分", "comment": "月度缺货退款率得分"},
    "monthlyNegativeReviewRateScore": {"name": "月度差评率得分", "comment": "月度差评率得分"},
    "monthlyPartialRefundRateScore": {"name": "月度部分退款率得分", "comment": "月度部分退款率得分"},
    "dailyMeituanRatingScore": {"name": "美团评分得分", "comment": "美团评分得分"},
    "dailyElemeRatingScore": {"name": "饿了么评分得分", "comment": "饿了么评分得分"},
    "monthlyMeituanDeliveryPunctualityRateScore": {"name": "美团配送准时率得分", "comment": "美团配送准时率得分"},
    "monthlyElemeTimelyDeliveryRateScore": {"name": "饿了么及时配送率得分", "comment": "饿了么及时配送率得分"},
    "validReplyWeightingPenalty": {"name": "有效回复权重惩罚", "comment": "有效回复权重惩罚"},
    "monthlyAverageStockRateWeightingPenalty": {"name": "月度平均库存率权重惩罚", "comment": "月度平均库存率权重惩罚"},
    "monthlyAverageTop500StockRateWeightingPenalty": {"name": "月度TOP500库存率权重惩罚", "comment": "月度TOP500库存率权重惩罚"},
    "monthlyAverageDirectStockRateWeightingPenalty": {"name": "月度直营库存率权重惩罚", "comment": "月度直营库存率权重惩罚"},
    "newProductComplianceListingWeightingPenalty": {"name": "新品合规上架权重惩罚", "comment": "新品合规上架权重惩罚"},
    "expiryManagementWeightingPenalty": {"name": "保质期管理权重惩罚", "comment": "保质期管理权重惩罚"},
    "inventoryLockWeightingPenalty": {"name": "库存锁定权重惩罚", "comment": "库存锁定权重惩罚"},
    "monthlyCumulativeHundredOrdersManhourWeightingPenalty": {"name": "月度累计百单工时权重惩罚", "comment": "月度累计百单工时权重惩罚"},
    "totalScoreWithoutWeightingPenalty": {"name": "无权重惩罚总分", "comment": "无权重惩罚总分"},
    "monthlyCumulativeMerchantLiabilityRefundRateWeightingPenalty": {"name": "月度累计商家责任退款率权重惩罚", "comment": "月度累计商家责任退款率权重惩罚"},
    "monthlyCumulativeOutOfStockRefundRateWeightingPenalty": {"name": "月度累计缺货退款率权重惩罚", "comment": "月度累计缺货退款率权重惩罚"},
    "meituanComplexExperienceScoreWeightingPenalty": {"name": "美团综合体验分权重惩罚", "comment": "美团综合体验分权重惩罚"},
    "meituanRatingWeightingPenalty": {"name": "美团评分权重惩罚", "comment": "美团评分权重惩罚"},
    "elemeRatingWeightingPenalty": {"name": "饿了么评分权重惩罚", "comment": "饿了么评分权重惩罚"},
    "partialRefundWeightingPenalty": {"name": "部分退款权重惩罚", "comment": "部分退款权重惩罚"},
    "trainingCompletedWeightingPenalty": {"name": "培训完成权重惩罚", "comment": "培训完成权重惩罚"},
    "totalWeightingPenalty": {"name": "总权重惩罚", "comment": "总权重惩罚"}
}

@app.get("/api/export-fields")
async def get_export_fields():
    """获取可导出的字段配置"""
    try:
        # 使用FIELD_CONFIG常量，确保与Excel导出功能一致
        # FIELD_CONFIG已经是Record格式，直接返回
        return {
            "success": True,
            "data": FIELD_CONFIG
        }
    except Exception as e:
        logger.error(f"获取导出字段失败: {e}")
        return {
            "success": False,
            "message": f"获取导出字段失败: {str(e)}"
        }

@app.get("/api/export-field-rule")
async def get_export_field_rule():
    """获取字段导出规则"""
    try:
        rule = config_db_manager.get_export_field_rule()
        return {
            "success": True,
            "data": rule
        }
    except Exception as e:
        logger.error(f"获取字段导出规则失败: {e}")
        return {
            "success": False,
            "message": f"获取字段导出规则失败: {str(e)}"
        }

@app.post("/api/export-field-rule")
async def save_export_field_rule(field_rule: dict = {}):
    """保存字段导出规则"""
    try:
        selected_fields = field_rule.get('selected_fields', [])
        config_db_manager.save_export_field_rule(selected_fields)
        return {
            "success": True,
            "message": "字段导出规则保存成功"
        }
    except Exception as e:
        logger.error(f"保存字段导出规则失败: {e}")
        return {
            "success": False,
            "message": f"保存字段导出规则失败: {str(e)}"
        }

@app.post("/api/export-excel")
async def export_excel(export_config: dict = {}):
    """导出数据到Excel文件"""
    try:
        # 获取导出规则
        rule = config_db_manager.get_export_rule()
        if rule:
            selected_stores = rule.get('selected_stores', [])
            rule_name = rule.get('rule_name', '默认导出规则')
        else:
            selected_stores = []
            rule_name = '全部数据'
        
        # 获取要导出的字段
        field_rule = config_db_manager.get_export_field_rule()
        if field_rule:
            selected_fields = field_rule.get('selected_fields', list(FIELD_CONFIG.keys()))
        else:
            selected_fields = export_config.get('selected_fields', list(FIELD_CONFIG.keys()))
        
        logger.info(f"选中的字段: {selected_fields}")
        
        # 获取日期参数
        export_date = export_config.get('date')
        
        # 获取数据库管理器
        db_manager = get_db_manager()
        
        # 查询数据
        data = []
        import re
        from datetime import datetime
        
        # 如果没有指定日期，默认查询今天
        if not export_date:
            export_date = datetime.now().strftime('%Y-%m-%d')
        
        logger.info(f"Excel导出查询日期: {export_date}")
        
        # 查询指定日期的数据
        query_date = datetime.strptime(export_date, '%Y-%m-%d')
        result = db_manager.query_data(date=query_date, page=1, page_size=10000)
        day_data = result.get('data', [])
        
        if selected_stores:
            # 根据选中的门店过滤数据
            store_names = []
            all_stores = config_db_manager.get_all_stores()
            store_id_to_name = {store['id']: store['store_name'] for store in all_stores}
            
            for store_id in selected_stores:
                if store_id in store_id_to_name:
                    store_names.append(store_id_to_name[store_id])
            
            if not store_names:
                return {
                    "success": False,
                    "message": "未找到有效的门店名称"
                }
            
            # 过滤选中的门店
            for item in day_data:
                mdshow = item.get('MDShow', '')
                clean_mdshow = re.sub(r'<[^>]+>', '', mdshow)
                if clean_mdshow in store_names:
                    data.append(item)
        else:
            # 导出所有数据
            data = day_data
        
        if not data:
            return {
                "success": False,
                "message": "没有找到数据"
            }
        
        # 按月份分组数据
        monthly_data = {}
        for item in data:
            create_date = item.get('createDateShow', '')
            if create_date:
                # 提取年月
                try:
                    date_obj = datetime.strptime(create_date, '%Y-%m-%d')
                    month_key = date_obj.strftime('%Y-%m')
                    if month_key not in monthly_data:
                        monthly_data[month_key] = []
                    monthly_data[month_key].append(item)
                except:
                    continue
        
        # 生成文件名（基于字段配置的哈希值，确保相同配置使用相同文件名）
        import hashlib
        fields_hash = hashlib.md5(''.join(sorted(selected_fields)).encode()).hexdigest()[:8]
        # 获取当前日期，格式化为YY_MM
        current_date = datetime.now().strftime("%y_%m")
        filename = f"ksx_{current_date}_{fields_hash}.xlsx"
        
        # 生成Excel内容（支持增量更新）
        excel_content = create_incremental_excel(monthly_data, selected_fields, rule_name, filename)
        
        logger.info(f"Excel数据导出成功: {filename}, 共 {len(data)} 条记录")
        
        return {
            "success": True,
            "message": f"Excel数据导出成功，共 {len(data)} 条记录",
            "filename": filename,
            "file_path": f"backend/exports/{filename}",
            "excel_content": excel_content.hex(),  # 转换为十六进制字符串
            "count": len(data)
        }
        
    except Exception as e:
        logger.error(f"导出Excel数据失败: {e}")
        return {
            "success": False,
            "message": f"导出Excel数据失败: {str(e)}"
        }

def create_incremental_excel(monthly_data: dict, selected_fields: List[str], rule_name: str, filename: str) -> bytes:
    """创建支持增量更新的Excel文件"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    import os
    
    # 使用绝对路径
    import os
    current_dir = os.path.dirname(os.path.abspath(__file__))
    exports_dir = os.path.join(current_dir, "exports")
    existing_file_path = os.path.join(exports_dir, filename)
    wb = None
    
    if os.path.exists(existing_file_path):
        try:
            # 加载现有文件
            wb = load_workbook(existing_file_path)
            logger.info(f"加载现有Excel文件: {existing_file_path}")
        except Exception as e:
            logger.warning(f"无法加载现有文件 {existing_file_path}: {e}")
            wb = None
    
    if wb is None:
        # 创建新的工作簿
        wb = Workbook()
        # 删除默认工作表
        wb.remove(wb.active)
        logger.info(f"创建新的Excel文件: {filename}")
    
    # 处理每个月的数据
    for month, month_data in monthly_data.items():
        sheet_name = f"{month}月"
        
        # 检查工作表是否存在
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"更新现有工作表: {sheet_name}")
        else:
            ws = wb.create_sheet(title=sheet_name)
            logger.info(f"创建新工作表: {sheet_name}")
        
        # 更新工作表数据
        update_monthly_sheet(ws, month_data, selected_fields, month)
    
    # 确保导出目录存在
    os.makedirs(exports_dir, exist_ok=True)
    logger.info(f"创建导出目录: {exports_dir}")
    
    # 保存文件
    wb.save(existing_file_path)
    logger.info(f"Excel文件已保存: {existing_file_path}")
    
    # 检查文件是否真的存在
    if os.path.exists(existing_file_path):
        logger.info(f"文件确认存在: {existing_file_path}")
        # 返回文件内容
        with open(existing_file_path, 'rb') as f:
            return f.read()
    else:
        logger.error(f"文件保存失败: {existing_file_path}")
        raise Exception(f"文件保存失败: {existing_file_path}")

def update_monthly_sheet(ws, data: List[Dict], selected_fields: List[str], month: str):
    """更新月度工作表数据 - 增量添加列"""
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    import re
    
    # 按门店和日期分组
    store_data = {}
    new_dates = set()
    
    for item in data:
        store_name = re.sub(r'<[^>]+>', '', item.get('MDShow', ''))
        create_date = item.get('createDateShow', '')
        
        if store_name and create_date:
            if store_name not in store_data:
                store_data[store_name] = {}
            store_data[store_name][create_date] = item
            new_dates.add(create_date)
    
    # 获取现有的列信息
    existing_columns = {}
    existing_stores = {}
    if ws.max_row > 0:
        # 读取第一行获取现有列
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and isinstance(cell_value, str):
                existing_columns[cell_value] = col
        
        # 读取第二列获取现有门店
        for row in range(2, ws.max_row + 1):
            store_name = ws.cell(row=row, column=2).value
            if store_name:
                existing_stores[store_name] = row
    
    # 检查是否有新的日期需要添加
    new_dates_to_add = []
    for date in new_dates:
        if date not in existing_columns:
            new_dates_to_add.append(date)
    
    # 如果没有新日期，直接返回
    if not new_dates_to_add:
        logger.info(f"所有日期都已存在，无需更新")
        return
    
    logger.info(f"需要添加的新日期: {new_dates_to_add}")
    
    # 获取所有现有日期和新日期，按顺序排列
    all_dates = sorted(list(existing_columns.keys()) + new_dates_to_add)
    
    # 获取所有门店（现有+新增）
    all_stores = sorted(set(list(existing_stores.keys()) + list(store_data.keys())))
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 检查是否需要设置基本表头
    need_header = False
    if ws.max_row == 0:
        need_header = True
    else:
        # 检查第1行第1列是否有"序号"
        first_cell = ws.cell(row=1, column=1).value
        if first_cell != "序号":
            need_header = True
    
    if need_header:
        # 第一列：序列号
        ws.cell(row=1, column=1, value="序号").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = header_alignment
        ws.cell(row=1, column=1).border = thin_border
        
        # 第二列：门店名称
        ws.cell(row=1, column=2, value="门店名称").font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        ws.cell(row=1, column=2).alignment = header_alignment
        ws.cell(row=1, column=2).border = thin_border
        
        # 第三列：指标名称
        ws.cell(row=1, column=3, value="指标名称").font = header_font
        ws.cell(row=1, column=3).fill = header_fill
        ws.cell(row=1, column=3).alignment = header_alignment
        ws.cell(row=1, column=3).border = thin_border
    
    # 添加新的日期列
    next_col = ws.max_column + 1 if ws.max_column > 0 else 3
    for date in sorted(new_dates_to_add):
        ws.cell(row=1, column=next_col, value=date).font = header_font
        ws.cell(row=1, column=next_col).fill = header_fill
        ws.cell(row=1, column=next_col).alignment = header_alignment
        ws.cell(row=1, column=next_col).border = thin_border
        existing_columns[date] = next_col
        next_col += 1
    
    # 填充数据
    if not existing_stores:
        # 如果是新文件，创建完整的表结构
        current_row = 2
        serial_number = 1
        for store_name in all_stores:
            store_start_row = current_row
            
            # 为每个字段创建行
            for field in selected_fields:
                field_name = FIELD_CONFIG.get(field, {}).get('name', field)
                
                # 设置序列号（合并单元格）
                ws.cell(row=store_start_row, column=1, value=serial_number)
                ws.cell(row=store_start_row, column=1).font = Font(bold=True)
                ws.cell(row=store_start_row, column=1).border = thin_border
                ws.cell(row=store_start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                
                # 设置门店名称（合并单元格）
                ws.cell(row=store_start_row, column=2, value=store_name)
                ws.cell(row=store_start_row, column=2).font = Font(bold=True)
                ws.cell(row=store_start_row, column=2).border = thin_border
                ws.cell(row=store_start_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
                
                # 设置指标名称
                ws.cell(row=store_start_row, column=3, value=field_name)
                ws.cell(row=store_start_row, column=3).border = thin_border
                ws.cell(row=store_start_row, column=3).alignment = Alignment(horizontal="left", vertical="center")
                
                # 填充所有日期的数据
                for date in all_dates:
                    if date in store_data.get(store_name, {}):
                        item = store_data[store_name][date]
                        value = item.get(field, '')
                        if field == 'id':
                            logger.info(f"ID字段处理: store_name={store_name}, date={date}, field={field}, value={value}, item_keys={list(item.keys())}")
                        col_index = existing_columns[date]
                        ws.cell(row=store_start_row, column=col_index, value=value)
                        ws.cell(row=store_start_row, column=col_index).border = thin_border
                        ws.cell(row=store_start_row, column=col_index).alignment = Alignment(horizontal="center", vertical="center")
                
                store_start_row += 1
            
            # 合并序列号和门店名称单元格
            merge_end_row = store_start_row - 1
            if merge_end_row > current_row:
                ws.merge_cells(f'A{current_row}:A{merge_end_row}')
                ws.merge_cells(f'B{current_row}:B{merge_end_row}')
            current_row = store_start_row
            serial_number += 1
    else:
        # 如果是现有文件，处理新日期和新门店
        # 1. 先处理现有门店的新日期数据
        for store_name in all_stores:
            store_row = existing_stores.get(store_name)
            if store_row is None:
                # 新门店，稍后处理
                continue
            
            # 为每个字段填充新日期的数据
            field_index = 0
            for field in selected_fields:
                row_index = store_row + field_index
                
                # 只填充新日期的数据
                for date in new_dates_to_add:
                    if date in store_data.get(store_name, {}):
                        item = store_data[store_name][date]
                        value = item.get(field, '')
                        col_index = existing_columns[date]
                        ws.cell(row=row_index, column=col_index, value=value)
                        ws.cell(row=row_index, column=col_index).border = thin_border
                        ws.cell(row=row_index, column=col_index).alignment = Alignment(horizontal="center", vertical="center")
                
                field_index += 1
        
        # 2. 处理新增门店（在文件底部添加）
        new_stores = [store for store in all_stores if store not in existing_stores]
        if new_stores:
            logger.info(f"发现新增门店: {new_stores}")
            current_row = ws.max_row + 1
            
            # 计算新增门店的起始序列号
            # 需要找到现有门店的最大序列号
            max_serial = 0
            for row in range(2, ws.max_row + 1):
                serial_cell = ws.cell(row=row, column=1).value
                if serial_cell and isinstance(serial_cell, (int, float)):
                    max_serial = max(max_serial, int(serial_cell))
            
            serial_number = max_serial + 1
            
            for store_name in new_stores:
                store_start_row = current_row
                
                # 为每个字段创建行
                for field in selected_fields:
                    field_name = FIELD_CONFIG.get(field, {}).get('name', field)
                    
                    # 设置序列号
                    ws.cell(row=store_start_row, column=1, value=serial_number)
                    ws.cell(row=store_start_row, column=1).font = Font(bold=True)
                    ws.cell(row=store_start_row, column=1).border = thin_border
                    ws.cell(row=store_start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                    
                    # 设置门店名称
                    ws.cell(row=store_start_row, column=2, value=store_name)
                    ws.cell(row=store_start_row, column=2).font = Font(bold=True)
                    ws.cell(row=store_start_row, column=2).border = thin_border
                    ws.cell(row=store_start_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
                    
                    # 设置指标名称
                    ws.cell(row=store_start_row, column=3, value=field_name)
                    ws.cell(row=store_start_row, column=3).border = thin_border
                    ws.cell(row=store_start_row, column=3).alignment = Alignment(horizontal="left", vertical="center")
                    
                    # 填充所有日期的数据
                    for date in all_dates:
                        if date in store_data.get(store_name, {}):
                            item = store_data[store_name][date]
                            value = item.get(field, '')
                            col_index = existing_columns[date]
                            ws.cell(row=store_start_row, column=col_index, value=value)
                            ws.cell(row=store_start_row, column=col_index).border = thin_border
                            ws.cell(row=store_start_row, column=col_index).alignment = Alignment(horizontal="center", vertical="center")
                    
                    store_start_row += 1
                
                # 合并序列号和门店名称单元格
                merge_end_row = store_start_row - 1
                if merge_end_row > current_row:
                    ws.merge_cells(f'A{current_row}:A{merge_end_row}')
                    ws.merge_cells(f'B{current_row}:B{merge_end_row}')
                current_row = store_start_row
                serial_number += 1
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8   # 序号
    ws.column_dimensions['B'].width = 25  # 门店名称
    ws.column_dimensions['C'].width = 20  # 指标名称
    for date in new_dates_to_add:
        col_letter = get_column_letter(existing_columns[date])
        ws.column_dimensions[col_letter].width = 12  # 日期列

def create_monthly_excel(data: List[Dict], selected_fields: List[str], month: str) -> bytes:
    """创建月度Excel数据，返回Excel文件的二进制数据"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    # 按门店和日期分组
    store_data = {}
    dates = set()
    
    for item in data:
        store_name = re.sub(r'<[^>]+>', '', item.get('MDShow', ''))
        create_date = item.get('createDateShow', '')
        
        if store_name and create_date:
            if store_name not in store_data:
                store_data[store_name] = {}
            store_data[store_name][create_date] = item
            dates.add(create_date)
    
    # 排序日期和门店
    sorted_dates = sorted(dates)
    sorted_stores = sorted(store_data.keys())
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}月"
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    store_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 创建表头
    current_row = 1
    
    # 第一行：序号 + 门店名称 + 指标名称 + 日期
    ws.cell(row=current_row, column=1, value="序号").font = header_font
    ws.cell(row=current_row, column=1).fill = header_fill
    ws.cell(row=current_row, column=1).border = border
    ws.cell(row=current_row, column=1).alignment = center_alignment
    
    ws.cell(row=current_row, column=2, value="门店名称").font = header_font
    ws.cell(row=current_row, column=2).fill = header_fill
    ws.cell(row=current_row, column=2).border = border
    ws.cell(row=current_row, column=2).alignment = center_alignment
    
    ws.cell(row=current_row, column=3, value="指标名称").font = header_font
    ws.cell(row=current_row, column=3).fill = header_fill
    ws.cell(row=current_row, column=3).border = border
    ws.cell(row=current_row, column=3).alignment = center_alignment
    
    # 添加日期列
    for col_idx, date in enumerate(sorted_dates, start=4):
        ws.cell(row=current_row, column=col_idx, value=date).font = header_font
        ws.cell(row=current_row, column=col_idx).fill = header_fill
        ws.cell(row=current_row, column=col_idx).border = border
        ws.cell(row=current_row, column=col_idx).alignment = center_alignment
    
    current_row += 1
    
    # 为每个门店创建数据行
    serial_number = 1
    for store_name in sorted_stores:
        store_start_row = current_row
        
        # 为每个字段创建一行
        for field_idx, field in enumerate(selected_fields):
            if field in FIELD_CONFIG:
                # 序列号（只在第一行显示）
                if field_idx == 0:
                    ws.cell(row=current_row, column=1, value=serial_number).fill = store_fill
                    ws.cell(row=current_row, column=1).border = border
                    ws.cell(row=current_row, column=1).alignment = center_alignment
                else:
                    ws.cell(row=current_row, column=1).fill = store_fill
                    ws.cell(row=current_row, column=1).border = border
                
                # 门店名称（只在第一行显示）
                if field_idx == 0:
                    ws.cell(row=current_row, column=2, value=store_name).fill = store_fill
                    ws.cell(row=current_row, column=2).border = border
                    ws.cell(row=current_row, column=2).alignment = center_alignment
                else:
                    ws.cell(row=current_row, column=2).fill = store_fill
                    ws.cell(row=current_row, column=2).border = border
                
                # 指标名称
                ws.cell(row=current_row, column=3, value=FIELD_CONFIG[field]['name']).border = border
                ws.cell(row=current_row, column=3).alignment = center_alignment
                
                # 添加每日数据
                for col_idx, date in enumerate(sorted_dates, start=4):
                    if date in store_data[store_name]:
                        value = store_data[store_name][date].get(field, '')
                        ws.cell(row=current_row, column=col_idx, value=value).border = border
                        ws.cell(row=current_row, column=col_idx).alignment = center_alignment
                    else:
                        ws.cell(row=current_row, column=col_idx, value='').border = border
                        ws.cell(row=current_row, column=col_idx).alignment = center_alignment
                
                current_row += 1
        
        # 合并序列号和门店名称单元格
        if current_row > store_start_row:
            ws.merge_cells(f'A{store_start_row}:A{current_row - 1}')
            ws.merge_cells(f'B{store_start_row}:B{current_row - 1}')
        
        serial_number += 1
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8   # 序号
    ws.column_dimensions['B'].width = 25  # 门店名称
    ws.column_dimensions['C'].width = 20  # 指标名称
    
    for col_idx in range(4, 4 + len(sorted_dates)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    # 保存到内存
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

# 异常处理
from fastapi.responses import JSONResponse

@app.exception_handler(404)
async def not_found_handler(request, exc):
    """404错误处理"""
    return JSONResponse(
        status_code=404,
        content={
            "success": False,
            "message": "接口不存在",
            "error_code": "NOT_FOUND"
        }
    )

@app.exception_handler(500)
async def internal_error_handler(request, exc):
    """500错误处理"""
    logger.error(f"服务器内部错误: {exc}")
    return JSONResponse(
        status_code=500,
        content={
            "success": False,
            "message": "服务器内部错误",
            "error_code": "INTERNAL_ERROR"
        }
    )

if __name__ == "__main__":
    import uvicorn
    
    # 创建logs目录
    os.makedirs("logs", exist_ok=True)
    
    logger.info("启动KSX数据查询API服务...")
    
    uvicorn.run(
        "main:app",
        host="127.0.0.1",
        port=18888,
        reload=True,
        log_level="info"
    )
